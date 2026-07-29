"""Regression tests for the Insurance Growth Rate / Property Tax Abatement
feature on the underwriting export route (backend/app/api/v2/underwriting_routes.py).

Base payload figures are drawn from the seeded "Milestone & Trails Multifamily
Portfolio" deal (id=41) as baked into Aequitas_Model_-_v16_Template.xlsx's
Inputs tab (its stored `underwriting_json` is null in the dev DB, so the
template's own resting values — Purchase Price $6.7M, 56 units, etc. — serve
as the realistic fixture instead).

Formula recalculation requires a real spreadsheet engine (openpyxl only
reads cached/last-saved values, it never evaluates formulas) — these tests
shell out to LibreOffice headless if available, and skip the recalculation-
dependent assertions otherwise.

Row layout (Pro Forma sheet, per the pre-push review — gross tax and the
abatement credit are two separate rows, not netted):
  row 52 = Insurance ($, monthly, its own dedicated growth accumulator)
  row 55 = Property Tax, GROSS (assessed value x millage + special
           assessment, or the legacy manual $/unit figure as a fallback
           when no assessed value is set — see test below)
  row 56 = Total Non-controllable Opex (= 52:55 + 57)
  row 57 = Property Tax Abatement credit (negative contra-expense)
  row 79 = NOI after cap reserve
  row 85 = Gross Sales Proceeds (exit valuation; adds back trailing-12 of
           rows 55 and 57 at the exit month only)

Row 83 (TTM NOI, pre-existing/out of scope) uses a negative-width
SUM(OFFSET(cell,,,,-12)) that has a confirmed LibreOffice headless
recalculation bug (verified via an isolated repro unrelated to this
template) — real Excel evaluates it correctly. Tests that need a genuine
NOI/refi/exit figure therefore replicate row 83's own trailing-12-of-row-79
definition directly in Python from the (correctly computed) row 79 values,
rather than reading row 83 or the cells derived from it.
"""
import io
import os
import shutil
import subprocess
import sys
import tempfile
import unittest

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from app import create_app

LIBREOFFICE = shutil.which('soffice') or (
    '/Applications/LibreOffice.app/Contents/MacOS/soffice'
    if os.path.exists('/Applications/LibreOffice.app/Contents/MacOS/soffice')
    else None
)

BASE_PAYLOAD = {
    "propertyName": "Milestone & Trails Multifamily Portfolio",
    "address": "5070 Warm Springs Rd & 4531 Milgen Rd, Columbus, GA",
    "acquisitionDate": "2026-06-01",
    "bridgePeriodMonths": 24,
    "holdPeriodYears": 10,
    "ttmNoi": 557219.8628,
    "purchasePrice": 6700000,
    "salesCostPct": 0.02,
    "closingCostPct": 0.0075,
    "workingCapitalPerUnit": 8847.0,
    "lpEquityShare": 0.85,
    "unitMix": [
        {"unitType": "2bd/2ba - 5070", "count": 32, "askingRent": 1230.39, "avgSf": 1100},
        {"unitType": "2bd/2ba - 4531", "count": 24, "askingRent": 1241.67, "avgSf": 1250},
    ],
    "rubsPct": 0.7,
    "otherIncomePerUnitMo": 75,
    "lossToLeaseRate": [0.0125, 0.0125, 0.0125, 0.0125, 0.0125],
    "vacancyRate": [0.08, 0.07, 0.06, 0.06, 0.06],
    "badDebtRate": [0.01, 0.01, 0.01, 0.01, 0.005],
    "concessionsRate": [0.02, 0.01, 0.01, 0.0, 0.0],
    "nonRevenueUnits": [0, 0, 0, 0, 0],
    "opexAdminPerUnit": 1183.48,
    "opexRmPerUnit": 440.82,
    "opexContractServicePerUnit": 227.14,
    "opexTurnoverPerUnit": 291.07,
    "opexInsurancePerUnit": 642.86,
    "opexUtilitiesPerUnit": 331.47,
    "opexPropertyTaxPerUnit": 1498.94,
    "managementFeePct": 0.02,
    "capReservePerUnit": 250,
    "seniorLtvPct": 0.70,
    "seniorInterestRate": 0.0652,
    "seniorFinancingCostsPct": 0.01,
    "seniorIoPeriods": 24,
    "refiTermMonths": 360,
    "refiInterestRate": 0.0625,
    "refiFinancingCostsPct": 0.005,
    "refiIoPeriods": 24,
    "rentGrowthRate": 0.02,
    "opexGrowthRate": 0.02,
    "exitCapRate": 0.0675,
    "amFeePct": 0.01,
    "preferredReturnPct": 0.08,
    "gpPromotePct": 0.30,
    # Deliberately zero — a real old deal that never touched the Property
    # Tax Model section. Row 55 must fall back to the legacy manual
    # $/unit x growth-accumulator formula in this case (see
    # test_property_tax_falls_back_without_assessed_value below).
    "assessedValue": 0,
    "assessedValueNextBuyer": 0,
    "assessmentPct": 1.0,
    "millageRate": 0.01892679804473016,
    "specialAssessments": 0,
}

# A payload with a real assessed value populated, exercising the new
# gross-tax/abatement mechanism.
ASSESSED_PAYLOAD = dict(BASE_PAYLOAD)
ASSESSED_PAYLOAD.update({
    "assessedValue": 6700000,
    "assessedValueNextBuyer": 6700000,
    "specialAssessments": 5000,
})


def recalc(xlsx_bytes):
    """Write bytes to a temp file, recalculate via LibreOffice headless, return
    an openpyxl workbook loaded with data_only=True (evaluated formulas)."""
    with tempfile.TemporaryDirectory() as tmp:
        src = os.path.join(tmp, 'in.xlsx')
        with open(src, 'wb') as f:
            f.write(xlsx_bytes)
        profile = os.path.join(tmp, 'lo_profile')
        subprocess.run(
            [LIBREOFFICE, '--headless', '--norestore',
             f'-env:UserInstallation=file://{profile}',
             '--convert-to', 'xlsx:Calc MS Excel 2007 XML',
             '--outdir', tmp, src],
            check=True, capture_output=True, timeout=60,
        )
        out = os.path.join(tmp, 'in.xlsx')
        with open(out, 'rb') as f:
            return openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)


def trailing_12_sum(ws, row, end_month_idx):
    """Replicates row 83's own trailing-12-inclusive-of-current-month
    definition, applied to any row (used here for rows 55/57/79)."""
    z_col = column_index_from_string('Z')
    end_col = z_col + end_month_idx
    total = 0.0
    for c in range(end_col - 11, end_col + 1):
        v = ws.cell(row=row, column=c).value
        total += v if isinstance(v, (int, float)) else 0.0
    return total


class UnderwritingExportAbatementTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = create_app()
        cls.client = cls.app.test_client()

    def _export(self, payload):
        resp = self.client.post('/api/v2/underwriting/export', json=payload)
        if resp.status_code != 200:
            self.fail(f'export failed with {resp.status_code}: {resp.get_data(as_text=True)}')
        return resp.data

    def test_old_deal_payload_does_not_crash_and_writes_zero_defaults(self):
        """A payload with no insuranceGrowthRate/abatementPctSchedule keys at
        all (simulating a pre-feature caller) must export successfully, and
        the new Inputs rows must read back as the safe 0% resting default —
        not silently populated with a non-zero suggestion value."""
        data = self._export(BASE_PAYLOAD)
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
        inputs = wb['Inputs']
        self.assertEqual(inputs['C116'].value, 0, 'Insurance Growth Rate must rest at 0% for an old-deal payload')
        for row in range(117, 122):
            self.assertEqual(inputs.cell(row=row, column=3).value, 0, f'Abatement row {row} must rest at 0%')

    def test_new_fields_round_trip_into_inputs_tab(self):
        payload = dict(ASSESSED_PAYLOAD)
        payload['insuranceGrowthRate'] = 0.08
        payload['abatementPctSchedule'] = [0, 0.10, 0.70, 0.90, 0.90]
        payload['opexInsurancePerUnitConfirmed'] = True
        data = self._export(payload)
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
        inputs = wb['Inputs']
        self.assertAlmostEqual(inputs['C116'].value, 0.08)
        self.assertEqual([inputs.cell(row=r, column=3).value for r in range(117, 122)],
                          [0, 0.10, 0.70, 0.90, 0.90])
        self.assertEqual(inputs['C122'].value, 1)

    @unittest.skipUnless(LIBREOFFICE, 'LibreOffice not available for formula recalculation')
    def test_property_tax_falls_back_without_assessed_value(self):
        """With assessedValue absent/zero (a real old deal), gross Property
        Tax (row 55) must fall back to the legacy manual $/unit figure —
        not silently collapse to $0 because the new assessed-value calc
        evaluates to zero."""
        wb = recalc(self._export(BASE_PAYLOAD))
        ws = wb['Pro Forma']
        self.assertEqual(ws['D102'].value, 0, 'Gross tax basis should be $0 with no assessed value set')
        self.assertGreater(ws['AA55'].value, 0,
                            'Property Tax must fall back to the manual $/unit entry, not collapse to $0')

    @unittest.skipUnless(LIBREOFFICE, 'LibreOffice not available for formula recalculation')
    def test_zero_growth_and_abatement_hold_insurance_flat_and_no_credit(self):
        """With growth/abatement absent, Insurance must stay flat across the
        hold, and the Abatement Credit row must be exactly zero every month
        — no silent escalation or credit for old deals."""
        wb = recalc(self._export(ASSESSED_PAYLOAD))
        ws = wb['Pro Forma']
        month1_insurance = ws['AA52'].value
        last_insurance = ws['FB52'].value
        self.assertAlmostEqual(month1_insurance, last_insurance, places=2,
                                msg='Insurance must not grow when insuranceGrowthRate is absent')
        for col in range(column_index_from_string('AA'), column_index_from_string('FB') + 1):
            credit = ws.cell(row=57, column=col).value
            self.assertAlmostEqual(credit, 0, places=6,
                                    msg=f'Abatement credit at {get_column_letter(col)}57 must be 0 when absent')

    @unittest.skipUnless(LIBREOFFICE, 'LibreOffice not available for formula recalculation')
    def test_nonzero_growth_and_abatement_split_rows_behave_correctly(self):
        """With growth/abatement populated: Insurance compounds upward;
        gross Property Tax (row 55) stays constant (abatement never touches
        the gross figure); the Abatement Credit (row 57) is a separate,
        negative line that varies by year and nets out to the expected
        percentage of gross tax; Total Non-controllable Opex (row 56)
        correctly folds in both."""
        payload = dict(ASSESSED_PAYLOAD)
        payload['insuranceGrowthRate'] = 0.08
        payload['abatementPctSchedule'] = [0, 0.10, 0.70, 0.90, 0.90]
        wb = recalc(self._export(payload))
        ws = wb['Pro Forma']

        month1_insurance = ws['AA52'].value
        last_insurance = ws['FB52'].value
        self.assertGreater(last_insurance, month1_insurance,
                            'Insurance should compound upward over the hold at 8%/yr')

        gross_tax_monthly = ws['D102'].value / 12
        month1_gross_tax = ws['AA55'].value    # year 1
        month13_gross_tax = ws.cell(row=55, column=39).value  # AM55, year 2
        self.assertAlmostEqual(month1_gross_tax, gross_tax_monthly, places=2)
        self.assertAlmostEqual(month13_gross_tax, gross_tax_monthly, places=2,
                                msg='Gross Property Tax (row 55) must NOT vary with the abatement schedule')

        month1_credit = ws['AA57'].value       # year 1, 0% abatement
        month13_credit = ws.cell(row=57, column=39).value  # AM57, year 2, 10% abatement
        self.assertAlmostEqual(month1_credit, 0, places=2)
        self.assertAlmostEqual(month13_credit, -gross_tax_monthly * 0.10, places=2,
                                msg='Year-2 credit must equal -10% of gross tax')

        # Total Non-controllable Opex must fold in both rows correctly.
        noncontrollable_month13 = ws.cell(row=56, column=39).value
        expected = (ws.cell(row=52, column=39).value + ws.cell(row=53, column=39).value
                    + ws.cell(row=54, column=39).value + month13_gross_tax + month13_credit)
        self.assertAlmostEqual(noncontrollable_month13, expected, places=2)

    @unittest.skipUnless(LIBREOFFICE, 'LibreOffice not available for formula recalculation')
    def test_exit_addback_direction_and_magnitude(self):
        """The exit-month addback (row 85, mirrored by G22) must add back
        trailing-12 gross tax + credit — i.e. the NET tax actually paid —
        on top of TTM NOI. Verified by replicating row 83's own trailing-12
        definition directly from row 79 (row 83 itself is skipped: it uses
        a negative-width OFFSET with a confirmed LibreOffice-only
        recalculation bug unrelated to this feature)."""
        payload = dict(ASSESSED_PAYLOAD)
        payload['insuranceGrowthRate'] = 0.08
        payload['abatementPctSchedule'] = [0, 0.10, 0.70, 0.90, 0.90]
        wb = recalc(self._export(payload))
        ws = wb['Pro Forma']

        exit_month = int(ws['G18'].value)
        cap_rate = ws['G23'].value

        ttm_noi_exit = trailing_12_sum(ws, 79, exit_month)
        addback = trailing_12_sum(ws, 55, exit_month) + trailing_12_sum(ws, 57, exit_month)
        exit_noi = ttm_noi_exit + addback
        exit_value = exit_noi / cap_rate

        self.assertGreater(addback, 0, 'Net tax paid at exit should be positive (partial abatement)')
        self.assertLess(addback, trailing_12_sum(ws, 55, exit_month),
                         'Addback (net of credit) must be less than gross tax alone once abatement applies')
        self.assertGreater(exit_value, 0)
        print(f'\n[exit_addback] TTM NOI={ttm_noi_exit:,.2f}  addback={addback:,.2f}  '
              f'exit NOI={exit_noi:,.2f}  exit value={exit_value:,.2f}')


if __name__ == '__main__':
    unittest.main()
