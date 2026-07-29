"""Cell-by-cell diff of two exported underwriting workbooks.

Dumps every non-empty cell's raw formula/value on the Inputs and Pro Forma
sheets of two .xlsx files and prints what differs. Used to verify that a
code change to the underwriting export route/template only touches the
cells it's supposed to (e.g. confirming an old-deal export is byte-for-byte
equivalent before/after adding a new, additive feature).

Usage:
    python3 diff_export_template.py <before.xlsx> <after.xlsx>
"""
import sys
import openpyxl

SHEETS = ('Inputs', 'Pro Forma')


def dump_cells(path):
    wb = openpyxl.load_workbook(path, data_only=False)
    out = {}
    for sheet in SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    out[(sheet, cell.coordinate)] = cell.value
    return out


def main():
    if len(sys.argv) != 3:
        print(f'Usage: python3 {sys.argv[0]} <before.xlsx> <after.xlsx>')
        sys.exit(1)

    before, after = dump_cells(sys.argv[1]), dump_cells(sys.argv[2])
    all_keys = set(before) | set(after)
    diffs = {k: (before.get(k), after.get(k)) for k in all_keys if before.get(k) != after.get(k)}

    for (sheet, coord), (b, a) in sorted(diffs.items()):
        print(f'{sheet}!{coord}  BEFORE: {b!r}  ->  AFTER: {a!r}')

    print(f'\n{len(diffs)} cell(s) differ.')


if __name__ == '__main__':
    main()
