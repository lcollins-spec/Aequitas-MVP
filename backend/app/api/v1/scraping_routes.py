"""
Property scraping API routes
Provides REST endpoints for extracting property data from listing URLs and PDFs
"""
import base64
import datetime
import io
import json
import os
import re
import tempfile
import uuid
from flask import Blueprint, request, jsonify
from werkzeug.utils import secure_filename
from app.services.scraping_service import ScrapingService
from app.services.pdf_extraction_service import PDFExtractionService
from app.database import db, PropertyImportModel, DealModel, DealDocumentModel

scraping_bp = Blueprint('scraping', __name__)


def _get_mime_type(filename: str) -> str:
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    return {
        'pdf': 'application/pdf',
        'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'xls': 'application/vnd.ms-excel',
        'csv': 'text/csv',
    }.get(ext, 'application/octet-stream')


def _save_document_to_drive(file_bytes: bytes, filename: str, deal_id_str: str, document_type: str) -> None:
    """Save file bytes to Google Drive and record in DB. Failures are logged but never raised."""
    try:
        deal = DealModel.query.get(int(deal_id_str))
        if not deal:
            return
        from app.utils import google_drive
        drive_file_id, drive_url = None, None
        try:
            result = google_drive.upload_file(
                file_bytes, filename, _get_mime_type(filename), deal.dealName, document_type
            )
            drive_file_id = result['file_id']
            drive_url = result['web_view_link']
            print(f'[{document_type}] Drive upload OK: {drive_url}', flush=True)
        except Exception as e:
            print(f'[{document_type}] Drive upload skipped: {e}', flush=True)
        doc = DealDocumentModel(
            id=str(uuid.uuid4()),
            deal_id=int(deal_id_str),
            file_name=filename,
            document_type=document_type,
            drive_file_id=drive_file_id,
            drive_url=drive_url,
            uploaded_at=datetime.datetime.utcnow(),
        )
        db.session.add(doc)
        db.session.commit()
        print(f'[{document_type}] DB record saved (deal_id={deal_id_str})', flush=True)
    except Exception as e:
        print(f'[{document_type}] Persistence skipped: {e}', flush=True)
        try:
            db.session.rollback()
        except Exception:
            pass


# Initialize services (lazy loading)
_scraping_service = None
_pdf_service = None


def get_scraping_service():
    """Get or create scraping service instance."""
    global _scraping_service
    if _scraping_service is None:
        _scraping_service = ScrapingService(cache_ttl=86400)  # 24 hours
    return _scraping_service


def get_pdf_service():
    """Get or create PDF extraction service instance."""
    global _pdf_service
    if _pdf_service is None:
        _pdf_service = PDFExtractionService()
    return _pdf_service


@scraping_bp.route('/scraping/extract', methods=['POST'])
def extract_property_data():
    """
    Extract property data from listing URL.

    Request Body:
        {
            "url": "https://www.loopnet.com/...",
            "enrichWithApi": true  // optional, default true
        }

    Returns:
        JSON response with extracted property data and import metadata
    """
    try:
        data = request.get_json()

        if not data:
            return jsonify({
                'success': False,
                'error': 'Request body is required',
                'code': 'INVALID_INPUT'
            }), 400

        url = data.get('url')
        if not url:
            return jsonify({
                'success': False,
                'error': 'URL is required',
                'code': 'INVALID_INPUT'
            }), 400

        enrich_with_api = data.get('enrichWithApi', True)

        # Extract property data
        scraping_service = get_scraping_service()
        result = scraping_service.extract_from_url(url, enrich=enrich_with_api)

        # Save import record to database
        import_record = PropertyImportModel(
            source_url=url,
            source_platform=result.source_platform,
            import_status=result.status,
            import_method=result.method,
            error_type=result.error_type,
            error_message=result.error_message,
            confidence_score=result.confidence_score,
            user_assisted=False
        )

        # Save extracted data as JSON
        if result.extracted_data:
            import_record.extracted_data = json.dumps(result.extracted_data.to_dict())

            # Also save individual fields for easy querying
            import_record.property_address = result.extracted_data.address
            import_record.city = result.extracted_data.city
            import_record.state = result.extracted_data.state
            import_record.zipcode = result.extracted_data.zipcode
            import_record.latitude = result.extracted_data.latitude
            import_record.longitude = result.extracted_data.longitude
            import_record.price = result.extracted_data.asking_price
            import_record.square_footage = result.extracted_data.building_size_sf
            import_record.units = result.extracted_data.num_units
            import_record.bedrooms = result.extracted_data.bedrooms
            import_record.bathrooms = result.extracted_data.bathrooms
            import_record.year_built = result.extracted_data.year_built
            import_record.property_type = result.extracted_data.property_type
            import_record.noi = result.extracted_data.noi
            import_record.cap_rate = result.extracted_data.cap_rate
            import_record.gross_income = result.extracted_data.gross_income

        # Save enrichment data
        if result.enrichment_data:
            import_record.enrichment_data = json.dumps(result.enrichment_data.to_dict())

        # Commit to database
        db.session.add(import_record)
        db.session.commit()

        # Return result
        if result.status == 'failed':
            return jsonify({
                'success': False,
                'error': result.error_message or 'Failed to extract property data',
                'code': 'EXTRACTION_FAILED',
                'details': {
                    'importId': import_record.id,
                    'status': result.status,
                    'errorType': result.error_type,
                    'errorMessage': result.error_message,
                    'suggestedAction': result.suggested_action
                }
            }), 400

        response_data = result.to_dict()
        response_data['importId'] = import_record.id

        return jsonify({
            'success': True,
            'data': response_data
        }), 200

    except ValueError as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'code': 'INVALID_INPUT'
        }), 400
    except Exception as e:
        print(f"Error in extract_property_data: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Internal server error',
            'code': 'SERVER_ERROR'
        }), 500


@scraping_bp.route('/scraping/extract-pdf', methods=['POST'])
def extract_from_pdf():
    """
    Extract property data from uploaded PDF file.

    Request:
        Multipart form data with 'file' field containing PDF

    Returns:
        JSON response with extracted property data and import metadata
    """
    try:
        # Check if file is present
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No file provided',
                'code': 'INVALID_INPUT'
            }), 400

        file = request.files['file']

        # Check if filename is empty
        if file.filename == '':
            return jsonify({
                'success': False,
                'error': 'No file selected',
                'code': 'INVALID_INPUT'
            }), 400

        # Check file extension
        if not file.filename.lower().endswith('.pdf'):
            return jsonify({
                'success': False,
                'error': 'File must be a PDF',
                'code': 'INVALID_FILE_TYPE'
            }), 400

        # Save file to temporary location
        filename = secure_filename(file.filename)
        temp_dir = tempfile.mkdtemp()
        temp_path = os.path.join(temp_dir, filename)

        try:
            file.save(temp_path)

            # Extract property data from PDF
            pdf_service = get_pdf_service()
            result = pdf_service.extract_from_pdf(temp_path)

            # Save import record to database
            import_record = PropertyImportModel(
                source_url=f'pdf_upload:{filename}',
                source_platform='pdf_upload',
                import_status=result.status,
                import_method=result.method,
                error_type=result.error_type,
                error_message=result.error_message,
                confidence_score=result.confidence_score,
                user_assisted=False
            )

            # Save extracted data as JSON
            if result.extracted_data:
                import_record.extracted_data = json.dumps(result.extracted_data.to_dict())

                # Also save individual fields for easy querying
                import_record.property_address = result.extracted_data.address
                import_record.city = result.extracted_data.city
                import_record.state = result.extracted_data.state
                import_record.zipcode = result.extracted_data.zipcode
                import_record.latitude = result.extracted_data.latitude
                import_record.longitude = result.extracted_data.longitude
                import_record.price = result.extracted_data.asking_price
                import_record.square_footage = result.extracted_data.building_size_sf
                import_record.units = result.extracted_data.num_units
                import_record.bedrooms = result.extracted_data.bedrooms
                import_record.bathrooms = result.extracted_data.bathrooms
                import_record.year_built = result.extracted_data.year_built
                import_record.property_type = result.extracted_data.property_type
                import_record.noi = result.extracted_data.noi
                import_record.cap_rate = result.extracted_data.cap_rate
                import_record.gross_income = result.extracted_data.gross_income

            # Commit to database
            db.session.add(import_record)
            db.session.commit()

            # Clean up temp file
            try:
                os.remove(temp_path)
                os.rmdir(temp_dir)
            except:
                pass  # Ignore cleanup errors

            # Return result
            if result.status == 'failed':
                return jsonify({
                    'success': False,
                    'error': result.error_message or 'Failed to extract property data from PDF',
                    'code': 'EXTRACTION_FAILED',
                    'details': {
                        'importId': import_record.id,
                        'status': result.status,
                        'errorType': result.error_type,
                        'errorMessage': result.error_message,
                        'suggestedAction': result.suggested_action
                    }
                }), 400

            response_data = result.to_dict()
            response_data['importId'] = import_record.id

            return jsonify({
                'success': True,
                'data': response_data
            }), 200

        finally:
            # Ensure cleanup even if an error occurs
            try:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                if os.path.exists(temp_dir):
                    os.rmdir(temp_dir)
            except:
                pass

    except Exception as e:
        print(f"Error in extract_from_pdf: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Internal server error',
            'code': 'SERVER_ERROR'
        }), 500


@scraping_bp.route('/scraping/imports/<int:import_id>', methods=['GET'])
def get_import(import_id):
    """
    Get a property import by ID.

    Path Parameters:
        import_id: ID of the import to retrieve

    Returns:
        JSON response with import data
    """
    try:
        import_record = PropertyImportModel.query.get(import_id)

        if not import_record:
            return jsonify({
                'success': False,
                'error': 'Import not found',
                'code': 'NOT_FOUND'
            }), 404

        # Parse JSON fields
        import_dict = import_record.to_dict()
        if import_record.extracted_data:
            try:
                import_dict['extractedData'] = json.loads(import_record.extracted_data)
            except:
                pass

        if import_record.enrichment_data:
            try:
                import_dict['enrichmentData'] = json.loads(import_record.enrichment_data)
            except:
                pass

        return jsonify({
            'success': True,
            'data': import_dict
        }), 200

    except Exception as e:
        print(f"Error in get_import: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Internal server error',
            'code': 'SERVER_ERROR'
        }), 500


@scraping_bp.route('/scraping/imports', methods=['GET'])
def list_imports():
    """
    List recent property imports with optional filters.

    Query Parameters:
        limit (optional): Maximum number of imports to return (default 20, max 100)
        status (optional): Filter by status ('success', 'partial', 'failed')
        dealId (optional): Filter by associated deal ID

    Returns:
        JSON response with imports array
    """
    try:
        limit = min(request.args.get('limit', 20, type=int), 100)
        status = request.args.get('status')
        deal_id = request.args.get('dealId', type=int)

        # Build query
        query = PropertyImportModel.query

        if status:
            query = query.filter_by(import_status=status)

        if deal_id:
            query = query.filter_by(deal_id=deal_id)

        # Order by most recent first
        query = query.order_by(PropertyImportModel.created_at.desc())

        # Apply limit
        imports = query.limit(limit).all()

        # Convert to dictionaries
        imports_data = []
        for import_record in imports:
            import_dict = import_record.to_dict()

            # Parse JSON fields
            if import_record.extracted_data:
                try:
                    import_dict['extractedData'] = json.loads(import_record.extracted_data)
                except:
                    pass

            if import_record.enrichment_data:
                try:
                    import_dict['enrichmentData'] = json.loads(import_record.enrichment_data)
                except:
                    pass

            imports_data.append(import_dict)

        return jsonify({
            'success': True,
            'data': {
                'imports': imports_data,
                'total': len(imports_data)
            }
        }), 200

    except Exception as e:
        print(f"Error in list_imports: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Internal server error',
            'code': 'SERVER_ERROR'
        }), 500


@scraping_bp.route('/scraping/imports/<int:import_id>', methods=['PATCH'])
def update_import(import_id):
    """
    Update a property import (e.g., link to deal, mark as user-assisted).

    Path Parameters:
        import_id: ID of the import to update

    Request Body:
        {
            "dealId": 123,  // optional
            "userAssisted": true  // optional
        }

    Returns:
        JSON response with updated import data
    """
    try:
        import_record = PropertyImportModel.query.get(import_id)

        if not import_record:
            return jsonify({
                'success': False,
                'error': 'Import not found',
                'code': 'NOT_FOUND'
            }), 404

        data = request.get_json()
        if not data:
            return jsonify({
                'success': False,
                'error': 'Request body is required',
                'code': 'INVALID_INPUT'
            }), 400

        # Update fields
        if 'dealId' in data:
            import_record.deal_id = data['dealId']

        if 'userAssisted' in data:
            import_record.user_assisted = data['userAssisted']

        # Commit changes
        db.session.commit()

        return jsonify({
            'success': True,
            'data': import_record.to_dict()
        }), 200

    except Exception as e:
        print(f"Error in update_import: {str(e)}")
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': 'Internal server error',
            'code': 'SERVER_ERROR'
        }), 500


@scraping_bp.route('/scraping/extract-om', methods=['POST'])
def extract_om():
    """
    Extract Offering Memorandum data from uploaded PDF using Claude's native document API.
    Extracts: property name, address, unit count, unit mix, and asking rents by unit type.
    Does not use any PDF parsing library — sends PDF directly to Claude as a base64 document block.
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided', 'code': 'INVALID_INPUT'}), 400

        file = request.files['file']

        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected', 'code': 'INVALID_INPUT'}), 400

        if not file.filename.lower().endswith('.pdf'):
            return jsonify({'success': False, 'error': 'File must be a PDF', 'code': 'INVALID_FILE_TYPE'}), 400

        api_key = os.getenv('ANTHROPIC_API_KEY')
        if not api_key:
            return jsonify({'success': False, 'error': 'ANTHROPIC_API_KEY not configured', 'code': 'SERVICE_UNAVAILABLE'}), 503

        try:
            from anthropic import Anthropic
        except ImportError:
            return jsonify({'success': False, 'error': 'Anthropic SDK not available', 'code': 'SERVICE_UNAVAILABLE'}), 503

        pdf_bytes = file.read()
        pdf_base64 = base64.b64encode(pdf_bytes).decode('utf-8')

        client = Anthropic(api_key=api_key)
        message = client.messages.create(
            model='claude-sonnet-4-6',
            max_tokens=4096,
            messages=[{
                'role': 'user',
                'content': [
                    {
                        'type': 'document',
                        'source': {
                            'type': 'base64',
                            'media_type': 'application/pdf',
                            'data': pdf_base64
                        }
                    },
                    {
                        'type': 'text',
                        'text': (
                            'Respond with raw JSON only. No markdown, no backticks, no explanation, no preamble. '
                            'Your entire response must be valid JSON starting with { and ending with }.\n\n'
                            'This is a multifamily real estate Offering Memorandum (OM). '
                            'Extract the following fields and return ONLY a valid JSON object with no extra text or markdown.\n\n'
                            'UNIT MIX RULES:\n'
                            '- Count units from the individual rent roll rows, NOT from summary tables.\n'
                            '- "count" = total units of that type including vacant (VACANT, V, empty rent) rows.\n'
                            '- "askingRent" = average monthly rent using ONLY occupied (non-vacant) rows for that type.\n'
                            '- Use unitType labels like: "Studio", "1BR/1BA", "2BR/1BA", "2BR/2BA", "3BR/2BA".\n'
                            '- Each entry in unitMix must represent ONE bedroom type.\n\n'
                            'OPERATING EXPENSES RULES:\n'
                            '- Pull annual totals from the operating expense summary table.\n'
                            '- If only a monthly figure is shown, multiply by 12 for annual.\n'
                            '- managementFeePct: express as a decimal between 0 and 1 (e.g., 0.06 for 6%).\n\n'
                            'VACANCY & CREDIT LOSS RULES:\n'
                            '- vacancyRate: look for "vacancy", "vacancy loss", "physical vacancy" in the income section. Express as decimal (e.g. 0.05 for 5%).\n'
                            '- badDebtRate: look for "bad debt", "credit loss", "collection loss". Express as decimal.\n'
                            '- lossToLeaseRate: look for "loss to lease" — the gap between market rent and in-place rent as a % of GPR. Express as decimal.\n'
                            '- concessionsRate: look for "concessions", "free rent", "lease-up concessions" as a % of GPR. Express as decimal.\n'
                            '- If only a dollar amount is shown, divide by GPR to get the rate.\n\n'
                            'OTHER INCOME RULES:\n'
                            '- parkingIncomePerUnit: monthly parking income per unit in dollars. If only a total annual figure is available, divide by 12 then by numUnits.\n'
                            '- rubsPct: RUBS or utility reimbursement percentage (0-100). Look for "RUBS", "utility reimbursement", "utility recovery", or similar. Return as a number 0-100, or null if not stated.\n'
                            '- otherIncomePerUnit: monthly per-unit income from sources other than rent, parking, and laundry (storage, pet fees, late fees, etc.). Exclude laundry and parking.\n'
                            '- laundryIncome: total ANNUAL laundry and vending income only (do not include parking or other misc income here).\n\n'
                            'PER-UNIT OPEX RULES:\n'
                            '- For each line item, return annual dollars per unit. If only a total annual amount is shown, divide by numUnits.\n'
                            '- opexPayrollPerUnit: payroll, salaries, benefits — annual $/unit.\n'
                            '- opexAdminPerUnit: administrative, office, legal, professional fees — annual $/unit.\n'
                            '- opexMarketingPerUnit: marketing, advertising, leasing — annual $/unit.\n'
                            '- opexRmPerUnit: repairs and maintenance only (not contract services) — annual $/unit.\n'
                            '- opexContractServicePerUnit: contract services, landscaping, janitorial, pest control — annual $/unit.\n'
                            '- opexTurnoverPerUnit: turnover, make-ready costs — annual $/unit.\n'
                            '- capexPerUnit: capital expenditure reserve or replacement reserve — annual $/unit.\n\n'
                            'GROWTH RATES RULES:\n'
                            '- opexGrowthRate: projected annual operating expense growth rate as a decimal. Look for "expense growth", "opex growth", or CPI-based assumptions.\n'
                            '- propertyTaxGrowthRate: projected annual property tax growth rate as a decimal. Look for tax escalation assumptions.\n'
                            '- If not stated explicitly, return null — do not fabricate a value.\n\n'
                            'RENT STABILIZATION RULES:\n'
                            '- Set rentStabilized to true if the document mentions rent stabilization, rent control, COLA caps, or CPI-linked rent increases.\n'
                            '- annualRentGrowthCap: the maximum allowed annual rent increase as a decimal (e.g., 0.022 for 2.2%).\n\n'
                            'Return this exact JSON structure (use null for any field not found):\n'
                            '{\n'
                            '  "propertyName": "string or null",\n'
                            '  "address": "full street address or null",\n'
                            '  "city": "string or null",\n'
                            '  "state": "2-letter state code or null",\n'
                            '  "zipcode": "string or null",\n'
                            '  "numUnits": integer or null,\n'
                            '  "askingPrice": number or null,\n'
                            '  "unitMix": [\n'
                            '    {"unitType": "e.g. 1BR/1BA", "count": integer, "sqft": integer or null, "askingRent": number or null}\n'
                            '  ],\n'
                            '  "laundryIncome": "annual laundry/vending income (not parking) as number or null",\n'
                            '  "vacancyRate": "vacancy loss as decimal 0.0-1.0 or null",\n'
                            '  "badDebtRate": "credit loss / bad debt as decimal 0.0-1.0 or null",\n'
                            '  "lossToLeaseRate": "loss to lease as decimal 0.0-1.0 or null",\n'
                            '  "concessionsRate": "concessions as decimal 0.0-1.0 or null",\n'
                            '  "parkingIncomePerUnit": "monthly parking $/unit or null",\n'
                            '  "rubsPct": "RUBS or utility reimbursement as percentage 0-100 or null",\n'
                            '  "otherIncomePerUnit": "monthly other income $/unit (excl. parking and laundry) or null",\n'
                            '  "operatingExpenses": {\n'
                            '    "utilitiesAnnual": "number or null",\n'
                            '    "insuranceAnnual": "number or null",\n'
                            '    "propertyTaxAnnual": "number or null",\n'
                            '    "repairsMaintenanceAnnual": "number or null",\n'
                            '    "managementFeePct": "decimal 0.0-1.0 or null"\n'
                            '  },\n'
                            '  "opexPayrollPerUnit": "annual payroll $/unit or null",\n'
                            '  "opexAdminPerUnit": "annual admin $/unit or null",\n'
                            '  "opexMarketingPerUnit": "annual marketing $/unit or null",\n'
                            '  "opexRmPerUnit": "annual R&M $/unit or null",\n'
                            '  "opexContractServicePerUnit": "annual contract services $/unit or null",\n'
                            '  "opexTurnoverPerUnit": "annual turnover $/unit or null",\n'
                            '  "capexPerUnit": "annual capex reserve $/unit or null",\n'
                            '  "opexGrowthRate": "annual opex growth as decimal or null",\n'
                            '  "propertyTaxGrowthRate": "annual property tax growth as decimal or null",\n'
                            '  "rentStabilized": "boolean or null",\n'
                            '  "annualRentGrowthCap": "decimal 0.0-1.0 or null"\n'
                            '}'
                        ),
                        'cache_control': {'type': 'ephemeral'}
                    }
                ]
            }]
        )

        print(f'[extract-om] stop_reason={message.stop_reason}, content blocks={len(message.content)}', flush=True)
        for i, block in enumerate(message.content):
            block_type = getattr(block, 'type', type(block).__name__)
            block_text = getattr(block, 'text', None)
            print(f'[extract-om] block[{i}] type={block_type} text_len={len(block_text) if block_text else "N/A"} preview={repr((block_text or "")[:200])}', flush=True)

        response_text = ''
        for block in message.content:
            if hasattr(block, 'text') and block.text:
                response_text = block.text.strip()
                break

        if not response_text:
            print(f'[extract-om] no text content in response', flush=True)
            return jsonify({'success': False, 'error': 'Claude returned an empty response', 'code': 'PARSE_ERROR'}), 500

        # Strip markdown fences if present
        if '```' in response_text:
            response_text = re.sub(r'^```json\s*', '', response_text)
            response_text = re.sub(r'^```\s*', '', response_text)
            response_text = re.sub(r'```$', '', response_text)
            response_text = response_text.strip()

        # Extract JSON object even if Claude prepended chain-of-thought text
        json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
        if not json_match:
            print(f'[extract-om] no JSON object found in response: {repr(response_text[:300])}', flush=True)
            return jsonify({'success': False, 'error': 'No JSON object found in Claude response', 'code': 'PARSE_ERROR'}), 500
        response_text = json_match.group(0)

        print(f'[extract-om] parsing JSON ({len(response_text)} chars): {repr(response_text[:300])}', flush=True)
        data = json.loads(response_text)
        print(f'[extract-om] extracted payload: {json.dumps(data, indent=2)}', flush=True)
        return jsonify({'success': True, 'data': data}), 200

    except json.JSONDecodeError as e:
        print(f'OM extraction JSON parse error: {str(e)}', flush=True)
        print(f'RAW CLAUDE RESPONSE: {repr(response_text)}', flush=True)
        try:
            with open('/tmp/claude_om_response.txt', 'w') as f:
                f.write(response_text)
            print('[extract-om] raw response written to /tmp/claude_om_response.txt', flush=True)
        except Exception:
            pass
        return jsonify({'success': False, 'error': f'Failed to parse Claude response as JSON: {str(e)}', 'code': 'PARSE_ERROR'}), 500
    except Exception as e:
        error_msg = str(e)
        print(f'Error in extract_om: {error_msg}', flush=True)
        # Surface Anthropic API errors (billing, rate limits, etc.) directly
        if 'credit balance is too low' in error_msg or 'billing' in error_msg.lower():
            return jsonify({'success': False, 'error': 'Anthropic API credits exhausted. Please add credits at console.anthropic.com/settings/billing.', 'code': 'BILLING_ERROR'}), 503
        if 'rate_limit' in error_msg.lower() or 'rate limit' in error_msg.lower():
            return jsonify({'success': False, 'error': 'Anthropic API rate limit reached. Please try again in a moment.', 'code': 'RATE_LIMIT'}), 429
        if 'invalid_api_key' in error_msg or 'authentication' in error_msg.lower():
            return jsonify({'success': False, 'error': 'Anthropic API key is invalid.', 'code': 'AUTH_ERROR'}), 503
        return jsonify({'success': False, 'error': 'Internal server error', 'code': 'SERVER_ERROR'}), 500


@scraping_bp.route('/scraping/debug-om', methods=['POST'])
def debug_om():
    """
    Debug route: same as /extract-om but returns the raw Claude response as plain text.
    Remove this route once the JSON parsing issue is resolved.
    """
    if 'file' not in request.files:
        return 'No file provided', 400

    file = request.files['file']
    if not file.filename.lower().endswith('.pdf'):
        return 'File must be a PDF', 400

    api_key = os.getenv('ANTHROPIC_API_KEY')
    if not api_key:
        return 'ANTHROPIC_API_KEY not configured', 503

    try:
        from anthropic import Anthropic
    except ImportError:
        return 'Anthropic SDK not available', 503

    pdf_bytes = file.read()
    pdf_base64 = base64.b64encode(pdf_bytes).decode('utf-8')

    client = Anthropic(api_key=api_key)
    message = client.messages.create(
        model='claude-sonnet-4-6',
        max_tokens=4096,
        messages=[{
            'role': 'user',
            'content': [
                {
                    'type': 'document',
                    'source': {
                        'type': 'base64',
                        'media_type': 'application/pdf',
                        'data': pdf_base64
                    }
                },
                {
                    'type': 'text',
                    'text': (
                        'Respond with raw JSON only. No markdown, no backticks, no explanation, no preamble. '
                        'Your entire response must be valid JSON starting with { and ending with }.\n\n'
                        'This is a multifamily real estate Offering Memorandum (OM). '
                        'Extract the following fields and return ONLY a valid JSON object with no extra text or markdown.\n\n'
                        'UNIT MIX RULES:\n'
                        '- Count units from the individual rent roll rows, NOT from summary tables.\n'
                        '- Exclude any rows marked as VACANT, vacant, or empty/V from counts.\n'
                        '- Group occupied units by bedroom type and calculate the average monthly asking/market rent for each type.\n'
                        '- Use unitType labels like: "Studio", "1BR/1BA", "2BR/1BA", "2BR/2BA", "3BR/2BA".\n'
                        '- Each entry in unitMix must represent ONE bedroom type with the total occupied count and average rent.\n\n'
                        'OPERATING EXPENSES RULES:\n'
                        '- Pull annual totals from the operating expense summary table.\n'
                        '- If only a monthly figure is shown, multiply by 12 for annual.\n'
                        '- managementFeePct: express as a decimal between 0 and 1 (e.g., 0.06 for 6%).\n\n'
                        'VACANCY & CREDIT LOSS RULES:\n'
                        '- vacancyRate: look for "vacancy", "vacancy loss", "physical vacancy" in the income section. Express as decimal (e.g. 0.05 for 5%).\n'
                        '- badDebtRate: look for "bad debt", "credit loss", "collection loss". Express as decimal.\n'
                        '- If only a dollar amount is shown, divide by GPR to get the rate.\n\n'
                        'RENT STABILIZATION RULES:\n'
                        '- Set rentStabilized to true if the document mentions rent stabilization, rent control, COLA caps, or CPI-linked rent increases.\n'
                        '- annualRentGrowthCap: the maximum allowed annual rent increase as a decimal (e.g., 0.022 for 2.2%).\n\n'
                        'Return this exact JSON structure (use null for any field not found):\n'
                        '{\n'
                        '  "propertyName": "string or null",\n'
                        '  "address": "full street address or null",\n'
                        '  "city": "string or null",\n'
                        '  "state": "2-letter state code or null",\n'
                        '  "zipcode": "string or null",\n'
                        '  "numUnits": integer or null,\n'
                        '  "askingPrice": number or null,\n'
                        '  "unitMix": [\n'
                        '    {"unitType": "e.g. 1BR/1BA", "count": integer, "sqft": integer or null, "askingRent": number or null}\n'
                        '  ],\n'
                        '  "laundryIncome": "annual laundry/vending/other income as number or null",\n'
                        '  "vacancyRate": "vacancy loss as decimal 0.0-1.0 or null",\n'
                        '  "badDebtRate": "credit loss / bad debt as decimal 0.0-1.0 or null",\n'
                        '  "operatingExpenses": {\n'
                        '    "utilitiesAnnual": "number or null",\n'
                        '    "insuranceAnnual": "number or null",\n'
                        '    "propertyTaxAnnual": "number or null",\n'
                        '    "repairsMaintenanceAnnual": "number or null",\n'
                        '    "managementFeePct": "decimal 0.0-1.0 or null"\n'
                        '  },\n'
                        '  "rentStabilized": "boolean or null",\n'
                        '  "annualRentGrowthCap": "decimal 0.0-1.0 or null"\n'
                        '}'
                    ),
                    'cache_control': {'type': 'ephemeral'}
                }
            ]
        }]
    )

    raw_text = ''
    for block in message.content:
        if hasattr(block, 'text') and block.text:
            raw_text = block.text
            break

    with open('/tmp/claude_om_response.txt', 'w') as f:
        f.write(raw_text)

    return raw_text, 200, {'Content-Type': 'text/plain; charset=utf-8'}


def _build_claude_content(file, prompt_text):
    """
    Build Claude message content blocks for a PDF or Excel/CSV file.
    - PDF: sent as a native base64 document block + a cached text prompt.
    - Excel/CSV: converted to plain-text via openpyxl/csv then sent as a text block.
    """
    filename_lower = file.filename.lower()

    if filename_lower.endswith('.pdf'):
        file_bytes = file.read()
        pdf_base64 = base64.b64encode(file_bytes).decode('utf-8')
        return [
            {
                'type': 'document',
                'source': {
                    'type': 'base64',
                    'media_type': 'application/pdf',
                    'data': pdf_base64,
                },
            },
            {
                'type': 'text',
                'text': prompt_text,
                'cache_control': {'type': 'ephemeral'},
            },
        ]
    else:
        # Excel / CSV: convert to tab-separated text
        import openpyxl

        file_bytes = file.read()
        text_parts = []

        if filename_lower.endswith('.csv'):
            import csv
            reader = csv.reader(io.StringIO(file_bytes.decode('utf-8', errors='replace')))
            for row in reader:
                line = '\t'.join(str(c) for c in row)
                if line.strip():
                    text_parts.append(line)
        else:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                text_parts.append(f'=== Sheet: {sheet_name} ===')
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else '' for c in row]
                    if any(c.strip() for c in cells):
                        text_parts.append('\t'.join(cells))

        doc_text = '\n'.join(text_parts)[:30000]
        return [
            {
                'type': 'text',
                'text': f'Document content:\n\n{doc_text}\n\n---\n\n{prompt_text}',
            },
        ]


def _call_claude_for_extraction(content, api_key):
    """Send content to Claude and return the parsed JSON dict."""
    from anthropic import Anthropic
    client = Anthropic(api_key=api_key)
    message = client.messages.create(
        model='claude-sonnet-4-6',
        max_tokens=4096,
        messages=[{'role': 'user', 'content': content}],
    )
    response_text = ''
    for block in message.content:
        if hasattr(block, 'text') and block.text:
            response_text = block.text.strip()
            break

    if not response_text:
        raise ValueError('Claude returned an empty response')

    if '```' in response_text:
        response_text = re.sub(r'^```json\s*', '', response_text)
        response_text = re.sub(r'^```\s*', '', response_text)
        response_text = re.sub(r'```$', '', response_text)
        response_text = response_text.strip()

    json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
    if not json_match:
        raise ValueError('No JSON object found in Claude response')

    return json.loads(json_match.group(0))


_VALID_DOC_EXTENSIONS = ('.pdf', '.xlsx', '.xls', '.csv')

_RENT_ROLL_PROMPT = (
    'Respond with raw JSON only. No markdown, no backticks, no explanation, no preamble. '
    'Your entire response must be valid JSON starting with { and ending with }.\n\n'
    'This is a multifamily real estate Rent Roll document. '
    'Extract unit mix and rental data and return ONLY a valid JSON object with no extra text or markdown.\n\n'
    'UNIT MIX RULES:\n'
    '- Count ALL units of each bedroom type, including vacant ones.\n'
    '- "count" = total units of that type (including vacant rows).\n'
    '- "askingRent" = average monthly rent using ONLY occupied (non-vacant) rows for that type.\n'
    '- Use unitType labels like: "Studio", "1BR/1BA", "2BR/1BA", "2BR/2BA", "3BR/2BA".\n'
    '- Each entry in unitMix must represent ONE bedroom type.\n\n'
    'VACANCY RULES:\n'
    '- vacancyRate: total vacant units / total units as a decimal (e.g. 0.05 for 5%).\n'
    '- badDebtRate: look for "bad debt", "credit loss", "collection loss" as a decimal.\n\n'
    'RENT STABILIZATION RULES:\n'
    '- Set rentStabilized to true if the document mentions rent stabilization, rent control, COLA caps, or CPI-linked rent increases.\n'
    '- annualRentGrowthCap: the maximum allowed annual rent increase as a decimal (e.g., 0.022 for 2.2%).\n\n'
    'Return this exact JSON structure (use null for any field not found):\n'
    '{\n'
    '  "propertyName": "string or null",\n'
    '  "address": "full street address or null",\n'
    '  "city": "string or null",\n'
    '  "state": "2-letter state code or null",\n'
    '  "zipcode": "string or null",\n'
    '  "numUnits": integer or null,\n'
    '  "askingPrice": null,\n'
    '  "unitMix": [\n'
    '    {"unitType": "e.g. 1BR/1BA", "count": integer, "sqft": integer or null, "askingRent": number or null}\n'
    '  ],\n'
    '  "laundryIncome": null,\n'
    '  "vacancyRate": "vacancy rate as decimal 0.0-1.0 or null",\n'
    '  "badDebtRate": "credit loss / bad debt as decimal 0.0-1.0 or null",\n'
    '  "operatingExpenses": null,\n'
    '  "rentStabilized": "boolean or null",\n'
    '  "annualRentGrowthCap": "decimal 0.0-1.0 or null"\n'
    '}'
)

_T12_PROMPT = (
    'Respond with raw JSON only. No markdown, no backticks, no explanation, no preamble. '
    'Your entire response must be valid JSON starting with { and ending with }.\n\n'
    'This is a multifamily real estate Trailing 12-Month (T12) Operating Statement. '
    'Extract operating income and expense data and return ONLY a valid JSON object.\n\n'
    'INCOME RULES:\n'
    '- laundryIncome: total annual laundry, vending, parking, or other miscellaneous income (number or null).\n'
    '- vacancyRate: annual vacancy loss as a decimal of GPR (e.g. 0.05 for 5%).\n'
    '- badDebtRate: bad debt / collection loss as a decimal of GPR (e.g. 0.02 for 2%).\n'
    '- If only a dollar amount is shown, divide by GPR to compute the rate.\n\n'
    'OPERATING EXPENSE RULES:\n'
    '- Pull annual totals. If only monthly figures are shown, multiply by 12.\n'
    '- managementFeePct: express as a decimal between 0 and 1 (e.g. 0.06 for 6%).\n\n'
    'RENT STABILIZATION RULES:\n'
    '- Set rentStabilized to true if the document mentions rent stabilization, rent control, COLA caps, or CPI-linked increases.\n'
    '- annualRentGrowthCap: the maximum allowed annual rent increase as a decimal (e.g., 0.022 for 2.2%).\n\n'
    'Return this exact JSON structure (use null for any field not found):\n'
    '{\n'
    '  "propertyName": "string or null",\n'
    '  "address": "full street address or null",\n'
    '  "city": "string or null",\n'
    '  "state": "2-letter state code or null",\n'
    '  "zipcode": "string or null",\n'
    '  "numUnits": integer or null,\n'
    '  "askingPrice": null,\n'
    '  "unitMix": [],\n'
    '  "laundryIncome": "annual miscellaneous/other income as number or null",\n'
    '  "vacancyRate": "vacancy loss as decimal 0.0-1.0 or null",\n'
    '  "badDebtRate": "credit loss / bad debt as decimal 0.0-1.0 or null",\n'
    '  "operatingExpenses": {\n'
    '    "utilitiesAnnual": "number or null",\n'
    '    "insuranceAnnual": "number or null",\n'
    '    "propertyTaxAnnual": "number or null",\n'
    '    "repairsMaintenanceAnnual": "number or null",\n'
    '    "managementFeePct": "decimal 0.0-1.0 or null"\n'
    '  },\n'
    '  "rentStabilized": "boolean or null",\n'
    '  "annualRentGrowthCap": "decimal 0.0-1.0 or null"\n'
    '}'
)


@scraping_bp.route('/scraping/extract-rent-roll', methods=['POST'])
def extract_rent_roll():
    """
    Extract unit mix and rent data from a Rent Roll PDF or Excel file.
    Accepts .pdf, .xlsx, .xls, .csv. Returns OmExtractedData-compatible JSON.
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided', 'code': 'INVALID_INPUT'}), 400

        file = request.files['file']

        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected', 'code': 'INVALID_INPUT'}), 400

        if not any(file.filename.lower().endswith(ext) for ext in _VALID_DOC_EXTENSIONS):
            return jsonify({'success': False, 'error': 'File must be a PDF, Excel, or CSV file', 'code': 'INVALID_FILE_TYPE'}), 400

        api_key = os.getenv('ANTHROPIC_API_KEY')
        if not api_key:
            return jsonify({'success': False, 'error': 'ANTHROPIC_API_KEY not configured', 'code': 'SERVICE_UNAVAILABLE'}), 503

        try:
            from anthropic import Anthropic  # noqa: F401 — verify SDK available
        except ImportError:
            return jsonify({'success': False, 'error': 'Anthropic SDK not available', 'code': 'SERVICE_UNAVAILABLE'}), 503

        file_bytes = file.read()
        file.seek(0)
        content = _build_claude_content(file, _RENT_ROLL_PROMPT)
        data = _call_claude_for_extraction(content, api_key)
        print(f'[extract-rent-roll] extracted: {json.dumps(data, indent=2)}', flush=True)
        deal_id = request.form.get('deal_id')
        if deal_id:
            _save_document_to_drive(file_bytes, file.filename, deal_id, 'Rent Roll')
        return jsonify({'success': True, 'data': data}), 200

    except json.JSONDecodeError as e:
        return jsonify({'success': False, 'error': f'Failed to parse Claude response as JSON: {str(e)}', 'code': 'PARSE_ERROR'}), 500
    except Exception as e:
        error_msg = str(e)
        print(f'Error in extract_rent_roll: {error_msg}', flush=True)
        if 'credit balance is too low' in error_msg or 'billing' in error_msg.lower():
            return jsonify({'success': False, 'error': 'Anthropic API credits exhausted. Please add credits at console.anthropic.com/settings/billing.', 'code': 'BILLING_ERROR'}), 503
        if 'rate_limit' in error_msg.lower() or 'rate limit' in error_msg.lower():
            return jsonify({'success': False, 'error': 'Anthropic API rate limit reached. Please try again in a moment.', 'code': 'RATE_LIMIT'}), 429
        if 'invalid_api_key' in error_msg or 'authentication' in error_msg.lower():
            return jsonify({'success': False, 'error': 'Anthropic API key is invalid.', 'code': 'AUTH_ERROR'}), 503
        return jsonify({'success': False, 'error': 'Internal server error', 'code': 'SERVER_ERROR'}), 500


@scraping_bp.route('/scraping/extract-t12', methods=['POST'])
def extract_t12():
    """
    Extract operating income and expenses from a T12 Operating Statement PDF or Excel file.
    Accepts .pdf, .xlsx, .xls, .csv. Returns OmExtractedData-compatible JSON.
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided', 'code': 'INVALID_INPUT'}), 400

        file = request.files['file']

        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected', 'code': 'INVALID_INPUT'}), 400

        if not any(file.filename.lower().endswith(ext) for ext in _VALID_DOC_EXTENSIONS):
            return jsonify({'success': False, 'error': 'File must be a PDF, Excel, or CSV file', 'code': 'INVALID_FILE_TYPE'}), 400

        api_key = os.getenv('ANTHROPIC_API_KEY')
        if not api_key:
            return jsonify({'success': False, 'error': 'ANTHROPIC_API_KEY not configured', 'code': 'SERVICE_UNAVAILABLE'}), 503

        try:
            from anthropic import Anthropic  # noqa: F401 — verify SDK available
        except ImportError:
            return jsonify({'success': False, 'error': 'Anthropic SDK not available', 'code': 'SERVICE_UNAVAILABLE'}), 503

        file_bytes = file.read()
        file.seek(0)
        content = _build_claude_content(file, _T12_PROMPT)
        data = _call_claude_for_extraction(content, api_key)
        print(f'[extract-t12] extracted: {json.dumps(data, indent=2)}', flush=True)
        deal_id = request.form.get('deal_id')
        if deal_id:
            _save_document_to_drive(file_bytes, file.filename, deal_id, 'T12')
        return jsonify({'success': True, 'data': data}), 200

    except json.JSONDecodeError as e:
        return jsonify({'success': False, 'error': f'Failed to parse Claude response as JSON: {str(e)}', 'code': 'PARSE_ERROR'}), 500
    except Exception as e:
        error_msg = str(e)
        print(f'Error in extract_t12: {error_msg}', flush=True)
        if 'credit balance is too low' in error_msg or 'billing' in error_msg.lower():
            return jsonify({'success': False, 'error': 'Anthropic API credits exhausted. Please add credits at console.anthropic.com/settings/billing.', 'code': 'BILLING_ERROR'}), 503
        if 'rate_limit' in error_msg.lower() or 'rate limit' in error_msg.lower():
            return jsonify({'success': False, 'error': 'Anthropic API rate limit reached. Please try again in a moment.', 'code': 'RATE_LIMIT'}), 429
        if 'invalid_api_key' in error_msg or 'authentication' in error_msg.lower():
            return jsonify({'success': False, 'error': 'Anthropic API key is invalid.', 'code': 'AUTH_ERROR'}), 503
        return jsonify({'success': False, 'error': 'Internal server error', 'code': 'SERVER_ERROR'}), 500


def _claude_pdf_or_text(file_bytes: bytes, filename: str, prompt: str, api_key: str) -> dict:
    """Send a file (PDF or spreadsheet) to Claude with a prompt and return parsed JSON."""
    import anthropic as _anthropic
    client = _anthropic.Anthropic(api_key=api_key)

    fname_lower = filename.lower()
    if fname_lower.endswith(('.xlsx', '.xls', '.csv')):
        # Convert spreadsheet to readable text
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            lines = []
            for ws in wb.worksheets:
                lines.append(f'--- Sheet: {ws.title} ---')
                for row in ws.iter_rows(values_only=True):
                    row_vals = [str(c) if c is not None else '' for c in row]
                    if any(v.strip() for v in row_vals):
                        lines.append('\t'.join(row_vals))
            text_content = '\n'.join(lines)
        except Exception:
            text_content = file_bytes.decode('utf-8', errors='replace')
        messages = [{'role': 'user', 'content': f'{prompt}\n\n{text_content}'}]
    else:
        b64 = base64.standard_b64encode(file_bytes).decode('utf-8')
        messages = [{
            'role': 'user',
            'content': [
                {
                    'type': 'document',
                    'source': {'type': 'base64', 'media_type': 'application/pdf', 'data': b64},
                },
                {'type': 'text', 'text': prompt},
            ],
        }]

    resp = client.messages.create(model='claude-opus-4-6', max_tokens=1024, messages=messages)
    text = resp.content[0].text.strip() if resp.content else '{}'
    # Strip markdown fences if present
    if text.startswith('```'):
        text = re.sub(r'^```[a-z]*\n?', '', text)
        text = re.sub(r'\n?```$', '', text)
    return json.loads(text)


_LOI_PROMPT = """Extract key terms from this Letter of Intent (LOI). Return ONLY a JSON object with these fields
(use null for any field not found):
{
  "purchasePrice": <number or null>,
  "earnestMoneyDeposit": <number or null>,
  "dueDiligenceDeadline": <"YYYY-MM-DD" or null>,
  "financingContingency": <"YYYY-MM-DD" or null>,
  "targetCloseDate": <"YYYY-MM-DD" or null>,
  "loanAmount": <number or null>,
  "interestRate": <number or null, as a percentage e.g. 6.5>,
  "loanTermMonths": <number or null>
}"""

_PSA_PROMPT = """Extract key terms from this Purchase and Sale Agreement (PSA). Return ONLY a JSON object with these fields
(use null for any field not found):
{
  "psaExecutedDate": <"YYYY-MM-DD" or null>,
  "earnestMoneyHardDate": <"YYYY-MM-DD" or null>,
  "purchasePrice": <number or null>,
  "closingDate": <"YYYY-MM-DD" or null>,
  "keyConditions": <string summarizing major contingencies or null>,
  "psaDraftedBy": <string or null>
}"""

_MODEL_PROMPT = """Extract key financial metrics from this real estate financial model (Excel or similar).
Return ONLY a JSON object with these fields (use null for any field not found):
{
  "purchasePrice": <number or null>,
  "totalEquityRequired": <number or null>,
  "acquisitionLoanAmount": <number or null>,
  "projectedLpNetIrr": <number or null, as a percentage e.g. 14.5>,
  "projectedEquityMultiple": <number or null, e.g. 1.85>,
  "projectedExitValue": <number or null>,
  "strategy": <string, e.g. "Value-Add" or null>
}"""


@scraping_bp.route('/scraping/extract-loi', methods=['POST'])
def extract_loi():
    """Extract deal terms from an LOI PDF."""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        file = request.files['file']
        api_key = os.getenv('ANTHROPIC_API_KEY')
        if not api_key:
            return jsonify({'success': False, 'error': 'ANTHROPIC_API_KEY not configured', 'code': 'SERVICE_UNAVAILABLE'}), 503
        file_bytes = file.read()
        data = _claude_pdf_or_text(file_bytes, file.filename or 'loi.pdf', _LOI_PROMPT, api_key)
        # Also save to Drive
        deal_id = request.form.get('deal_id')
        if deal_id:
            _save_document_to_drive(file_bytes, file.filename, deal_id, 'LOI Draft')
        return jsonify({'success': True, 'data': data}), 200
    except json.JSONDecodeError as e:
        return jsonify({'success': False, 'error': f'Failed to parse Claude response: {e}', 'code': 'PARSE_ERROR'}), 500
    except Exception as e:
        print(f'Error in extract_loi: {e}', flush=True)
        return jsonify({'success': False, 'error': str(e), 'code': 'SERVER_ERROR'}), 500


@scraping_bp.route('/scraping/extract-psa', methods=['POST'])
def extract_psa():
    """Extract deal terms from a PSA PDF."""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        file = request.files['file']
        api_key = os.getenv('ANTHROPIC_API_KEY')
        if not api_key:
            return jsonify({'success': False, 'error': 'ANTHROPIC_API_KEY not configured', 'code': 'SERVICE_UNAVAILABLE'}), 503
        file_bytes = file.read()
        data = _claude_pdf_or_text(file_bytes, file.filename or 'psa.pdf', _PSA_PROMPT, api_key)
        deal_id = request.form.get('deal_id')
        if deal_id:
            _save_document_to_drive(file_bytes, file.filename, deal_id, 'PSA Draft')
        return jsonify({'success': True, 'data': data}), 200
    except json.JSONDecodeError as e:
        return jsonify({'success': False, 'error': f'Failed to parse Claude response: {e}', 'code': 'PARSE_ERROR'}), 500
    except Exception as e:
        print(f'Error in extract_psa: {e}', flush=True)
        return jsonify({'success': False, 'error': str(e), 'code': 'SERVER_ERROR'}), 500


@scraping_bp.route('/scraping/extract-model', methods=['POST'])
def extract_model():
    """Extract financial metrics from an Excel financial model."""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        file = request.files['file']
        api_key = os.getenv('ANTHROPIC_API_KEY')
        if not api_key:
            return jsonify({'success': False, 'error': 'ANTHROPIC_API_KEY not configured', 'code': 'SERVICE_UNAVAILABLE'}), 503
        file_bytes = file.read()
        file.seek(0)
        data = _claude_pdf_or_text(file_bytes, file.filename or 'model.xlsx', _MODEL_PROMPT, api_key)
        deal_id = request.form.get('deal_id')
        if deal_id:
            _save_document_to_drive(file_bytes, file.filename, deal_id, 'Financial Model')
        return jsonify({'success': True, 'data': data}), 200
    except json.JSONDecodeError as e:
        return jsonify({'success': False, 'error': f'Failed to parse Claude response: {e}', 'code': 'PARSE_ERROR'}), 500
    except Exception as e:
        print(f'Error in extract_model: {e}', flush=True)
        return jsonify({'success': False, 'error': str(e), 'code': 'SERVER_ERROR'}), 500


@scraping_bp.route('/scraping/summarize-dd-doc', methods=['POST'])
def summarize_dd_doc():
    """Summarize a DD document in 2-3 sentences with key findings."""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        file = request.files['file']
        api_key = os.getenv('ANTHROPIC_API_KEY')
        if not api_key:
            return jsonify({'success': False, 'error': 'ANTHROPIC_API_KEY not configured', 'code': 'SERVICE_UNAVAILABLE'}), 503
        file_bytes = file.read()

        import anthropic as _anthropic
        client = _anthropic.Anthropic(api_key=api_key)
        fname_lower = (file.filename or '').lower()
        if fname_lower.endswith(('.xlsx', '.xls', '.csv')):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                lines = []
                for ws in wb.worksheets:
                    lines.append(f'--- Sheet: {ws.title} ---')
                    for row in ws.iter_rows(values_only=True):
                        row_vals = [str(c) if c is not None else '' for c in row]
                        if any(v.strip() for v in row_vals):
                            lines.append('\t'.join(row_vals))
                text_content = '\n'.join(lines)
            except Exception:
                text_content = file_bytes.decode('utf-8', errors='replace')
            messages = [{
                'role': 'user',
                'content': (
                    'Summarize this due diligence document in 2-3 sentences. '
                    f'Identify any red flags, required actions, or key findings.\n\n{text_content}'
                ),
            }]
        else:
            b64 = base64.standard_b64encode(file_bytes).decode('utf-8')
            messages = [{
                'role': 'user',
                'content': [
                    {
                        'type': 'document',
                        'source': {'type': 'base64', 'media_type': 'application/pdf', 'data': b64},
                    },
                    {
                        'type': 'text',
                        'text': (
                            'Summarize this due diligence document in 2-3 sentences. '
                            'Identify any red flags, required actions, or key findings.'
                        ),
                    },
                ],
            }]

        resp = client.messages.create(model='claude-opus-4-6', max_tokens=512, messages=messages)
        summary = resp.content[0].text.strip() if resp.content else ''
        return jsonify({'success': True, 'summary': summary}), 200
    except Exception as e:
        print(f'Error in summarize_dd_doc: {e}', flush=True)
        return jsonify({'success': False, 'error': str(e), 'code': 'SERVER_ERROR'}), 500


# Error handlers
@scraping_bp.errorhandler(404)
def not_found(error):
    """Handle 404 errors."""
    return jsonify({
        'success': False,
        'error': 'Resource not found',
        'code': 'NOT_FOUND'
    }), 404


@scraping_bp.errorhandler(500)
def internal_error(error):
    """Handle 500 errors."""
    return jsonify({
        'success': False,
        'error': 'Internal server error',
        'code': 'SERVER_ERROR'
    }), 500
