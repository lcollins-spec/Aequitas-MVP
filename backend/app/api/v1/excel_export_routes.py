"""
Excel Export API Routes
Writes deal inputs into the v13 template and returns the file.
All formula cells are left untouched — Excel evaluates them on open.
"""
from flask import Blueprint, request, jsonify, send_file
from datetime import datetime, date
from io import BytesIO
import os

import openpyxl
from openpyxl import load_workbook
from app.database import DealModel

excel_export_bp = Blueprint('excel_export', __name__)

TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__), '..', '..', '..', 'MF_Acq_Pro_Forma_Template_-_v13.xlsx'
)

YEAR_COLS = ('D', 'E', 'F', 'G', 'H')


def _to_decimal(value, fallback=0.0):
    """Normalize a value to a decimal fraction.

    Values > 1 are assumed to be percentages (e.g. 65 → 0.65).
    Values ≤ 1 are returned as-is.
    None returns fallback.
    """
    if value is None:
        return float(fallback)
    v = float(value)
    return v / 100.0 if v > 1.0 else v


def _write_5yr(ws, row, value_or_list):
    """Write a value (or list of up to 5 values) across columns D–H for row.

    A scalar is replicated to all 5 columns.
    A short list is extended by repeating its last element.
    """
    if isinstance(value_or_list, list):
        last = value_or_list[-1] if value_or_list else 0
        vals = (list(value_or_list) + [last] * 5)[:5]
    else:
        vals = [value_or_list] * 5
    for col, val in zip(YEAR_COLS, vals):
        ws[f'{col}{row}'] = val


@excel_export_bp.route('/underwriting/<int:deal_id>/export-excel', methods=['POST'])
def export_underwriting_excel(deal_id):
    try:
        return _do_export(deal_id)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def _do_export(deal_id):
    data = request.get_json(silent=True) or {}

    deal = DealModel.query.get(deal_id)
    if deal is None:
        return jsonify({'error': f'Deal {deal_id} not found'}), 404

    if not os.path.exists(TEMPLATE_PATH):
        return jsonify({'error': f'Excel template not found at: {os.path.abspath(TEMPLATE_PATH)}'}), 500

    with open(TEMPLATE_PATH, 'rb') as f:
        template_bytes = BytesIO(f.read())
    wb = load_workbook(template_bytes)
    ws = wb.worksheets[0]

    # ── Property Info ─────────────────────────────────────────────────────────
    ws['D5'] = data.get('propertyName') or deal.deal_name or ''
    ws['D6'] = data.get('address') or deal.property_address or ''

    # ── Timeline ──────────────────────────────────────────────────────────────
    acq_date = None
    acq_date_raw = data.get('acquisitionDate')
    if acq_date_raw:
        try:
            acq_date = datetime.strptime(acq_date_raw[:10], '%Y-%m-%d').date()
        except (ValueError, TypeError):
            pass
    ws['E16'] = acq_date or date.today()

    hold_months = data.get('holdPeriodMonths')
    if hold_months is None:
        hold_months = int((data.get('holdPeriodYears') or 0) * 12)
    hold_years = int(hold_months) // 12

    # F17 is the senior loan duration — template pays off senior debt at this month.
    # Must not exceed hold period or the levered model breaks.
    bridge_loan_term_months = int(data.get('bridgeLoanTermMonths') or 36)
    ws['F17'] = min(bridge_loan_term_months, int(hold_months))  # capped at hold period
    ws['C19'] = hold_years            # exit year (hold period)
    ws['D48'] = '=F17'               # senior loan term also drives D48 amortization anchor

    # ── Valuation ─────────────────────────────────────────────────────────────
    ws['D23'] = float(data.get('ttmNoi') or 0)
    ws['D25'] = float(data.get('purchasePrice') or deal.purchase_price or 0)
    ws['E24'] = _to_decimal(data.get('entryCapRate') or 0)
    ws['G24'] = _to_decimal(data.get('refiCapRate') or 0)
    ws['H24'] = _to_decimal(data.get('exitCapRate') or 0)
    ws['E86'] = '=E25'  # sources-and-uses price ties to valuation price

    # ── Unit Mix (always 4 rows; zero-pad unused rows) ─────────────────────────
    unit_mix = data.get('unitMix') or []
    for i, row_num in enumerate([60, 61, 62, 63]):
        if i < len(unit_mix):
            u = unit_mix[i]
            ws[f'D{row_num}'] = int(u.get('count') or 0)
            ws[f'E{row_num}'] = float(u.get('askingRent') or u.get('rent') or 0)
            ws[f'F{row_num}'] = float(u.get('avgSf') or u.get('sf') or 0)
        else:
            ws[f'D{row_num}'] = 0
            ws[f'E{row_num}'] = 0
            ws[f'F{row_num}'] = 0

    # ── Other Income ──────────────────────────────────────────────────────────
    ws['E67'] = _to_decimal(data.get('rubsPct') or 0)
    ws['G68'] = float(data.get('parkingIncomePerUnit') or 0)
    ws['G69'] = float(data.get('otherIncomePerUnit') or 0)

    # ── Income Adjustments (rows 73–77, columns D–H = years 1–5) ──────────────
    def _pct_or_list(key):
        v = data.get(key)
        if isinstance(v, list):
            return [_to_decimal(x) for x in v]
        return _to_decimal(v)

    _write_5yr(ws, 73, _pct_or_list('lossToLeaseRate'))
    _write_5yr(ws, 74, _pct_or_list('vacancyRate'))
    _write_5yr(ws, 75, _pct_or_list('badDebtRate'))
    _write_5yr(ws, 76, _pct_or_list('concessionsRate'))
    _write_5yr(ws, 77, int(data.get('nonRevenueUnits') or 0))

    # ── Senior Financing ──────────────────────────────────────────────────────
    ws['D47'] = _to_decimal(data.get('ltv') or 0)
    ws['D50'] = _to_decimal(data.get('financingCostsPct') or 0.01)
    ws['E49'] = _to_decimal(data.get('interestRate') or 0)
    ws['D51'] = int(data.get('seniorIoPeriods') or 0)
    ws['C39'] = _to_decimal(data.get('closingCostsPct') or 0)

    # ── Refi Assumptions ──────────────────────────────────────────────────────
    refi_term = int(data.get('refiTermMonths') or 0)
    ws['G48'] = refi_term if refi_term > 0 else 360
    ws['G50'] = _to_decimal(data.get('refiFinancingCostsPct') or 0.005)
    ws['G51'] = int(data.get('refiIoPeriods') or 24)
    ws['G54'] = _to_decimal(data.get('refiDebtYield') or 0)
    ws['G55'] = _to_decimal(data.get('refiLtv') or 0.75)

    # ── Operating Expenses ($/unit/year) ──────────────────────────────────────
    ws['K42'] = float(data.get('opexPayrollPerUnit') or 0)
    ws['K43'] = float(data.get('opexAdminPerUnit') or 0)
    ws['K44'] = float(data.get('opexMarketingPerUnit') or 0)
    ws['K45'] = float(data.get('opexRmPerUnit') or 0)
    ws['K46'] = float(data.get('opexContractServicePerUnit') or 0)
    ws['K47'] = float(data.get('opexTurnoverPerUnit') or 0)
    ws['K48'] = float(data.get('opexOtherPerUnit') or 0)
    ws['K52'] = float(data.get('opexInsurancePerUnit') or 0)
    ws['K53'] = float(data.get('opexUtilitiesPerUnit') or 0)
    ws['K64'] = float(data.get('capexPerUnit') or 0)

    # ── Management Fee ────────────────────────────────────────────────────────
    ws['J54'] = _to_decimal(data.get('managementFeePct') or 0)

    # ── Property Tax ──────────────────────────────────────────────────────────
    millage = _to_decimal(data.get('millageRate') or 0)
    ws['E89'] = millage
    ws['F89'] = millage
    ws['E90'] = float(data.get('specialAssessments') or 0)
    ws['K55'] = '=E92'  # property tax $/unit pulls from template's computed total

    # ── Growth Rates ──────────────────────────────────────────────────────────
    ws['D79'] = _to_decimal(data.get('rentGrowthRate') or 0)
    ws['D80'] = _to_decimal(data.get('rentGrowthRate') or 0)
    ws['D81'] = _to_decimal(data.get('opexGrowthRate') or 0)
    ws['D82'] = _to_decimal(data.get('propertyTaxGrowthRate') or 0)
    ws['D105'] = _to_decimal(data.get('generalInflationRate') or data.get('rentGrowthRate') or 0)

    # ── GP Equity Share ───────────────────────────────────────────────────────
    ws['G8'] = 0.5

    # ── Dynamic IRR/SUM formula ranges ────────────────────────────────────────
    # Column V = index 22. exit_col = column at V + hold_period_months.
    hold_period_months = int(data.get('holdPeriodMonths') or 120)
    exit_col = openpyxl.utils.get_column_letter(22 + hold_period_months)

    ws['M88']  = f'=(1+IRR(V85:{exit_col}85,0))^12-1'
    ws['M105'] = f'=(1+IRR(V101:{exit_col}101,0))^12-1'
    ws['M118'] = f'=(1+IRR(V115:{exit_col}115,0))^12-1'
    ws['M127'] = f'=(1+IRR(V124:{exit_col}124,0))^12-1'

    ws['M87']  = f'=SUMIF(V85:{exit_col}85,"<0",V85:{exit_col}85)'
    ws['M104'] = f'=SUMIF(V101:{exit_col}101,"<0",V101:{exit_col}101)'
    ws['M116'] = f'=SUM(V115:{exit_col}115)'
    ws['M117'] = f'=SUMIF(V115:{exit_col}115,"<0",V115:{exit_col}115)'
    ws['M125'] = f'=SUM(V124:{exit_col}124)'
    ws['M126'] = f'=SUMIF(V124:{exit_col}124,"<0",V124:{exit_col}124)'

    ws['M170'] = f'=SUM(V170:{exit_col}170)'
    ws['M171'] = f'=SUM(V171:{exit_col}171)'
    ws['M172'] = f'=SUM(V172:{exit_col}172)'
    ws['M173'] = f'=SUM(V173:{exit_col}173)'
    ws['M174'] = f'=SUM(V174:{exit_col}174)'

    # ── Save and return ───────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    property_name = (data.get('propertyName') or deal.deal_name or 'Property').replace(' ', '_')
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{property_name}_ProForma.xlsx',
    )


@excel_export_bp.route('/underwriting/export-excel-template', methods=['GET'])
def export_template():
    try:
        timestamp = datetime.now().strftime('%Y%m%d')
        return send_file(
            TEMPLATE_PATH,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'MF_ProForma_Template_{timestamp}.xlsx',
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500
