import os
import io
import re
import json
import base64
import logging
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file

underwriting_v2_bp = Blueprint('underwriting_v2', __name__)

logger = logging.getLogger(__name__)

TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))),
    'Aequitas_Model_-_v16_Template.xlsx',
)

# ─── Inputs-tab row map ─────────────────────────────────────────────────────
# Single source of truth for where each field lands on the "Inputs" sheet.
# Adding a new template field should only ever require a new entry here.
ROWS = {
    # Property
    'property_name': 81,
    'address': 82,
    # Timeline
    'acquisition_date': 5,           # drives exit date formula on Pro Forma
    'bridge_term_months': 7,         # senior/bridge loan term in months
    'hold_period_years': 83,
    # Valuation
    'ttm_noi': 8,
    'purchase_price': 84,
    'cap_rate_valuation': 9,         # feeds the equity-basis formula on Pro Forma
    'sales_cost_pct': 10,
    'working_capital_reserve': 85,
    # LP / GP split
    'lp_equity_share': 86,
    # Unit mix (16 rows, cols B/C/D/E = type/count/rent/avgSf)
    'unit_mix_start': 96,
    'unit_mix_rows': 16,
    # Other income
    'rubs_pct': 87,
    'parking_per_unit': 88,
    'other_income_per_unit': 89,
    # Income adjustments (5-yr, one Inputs row per year)
    'loss_to_lease_rows': [24, 29, 34, 39, 44],
    'vacancy_rows': [25, 30, 35, 40, 45],
    'bad_debt_rows': [26, 31, 36, 41, 46],
    'concessions_rows': [27, 32, 37, 42, 47],
    'non_revenue_units_rows': [28, 33, 38, 43, 48],
    # Operating expenses ($/unit/yr)
    'opex_payroll_per_unit': 72,
    'opex_admin_per_unit': 73,
    'opex_marketing_per_unit': 74,
    'opex_rm_per_unit': 75,
    'opex_contract_service_per_unit': 76,
    'opex_turnover_per_unit': 77,
    'opex_other_per_unit': 78,
    'opex_insurance_per_unit': 60,
    'opex_utilities_per_unit': 61,
    'opex_property_tax_per_unit': 63,
    'management_fee_pct': 62,
    'cap_reserve_per_unit': 58,
    # Senior / bridge loan
    'senior_loan_amount': 67,
    'senior_interest_rate': 68,
    'senior_financing_costs_pct': 79,
    'senior_io_periods': 80,
    # Refinance loan
    # Note: DSCR/DY/LTV sizing targets (F52/F53/F54 on Pro Forma) are fixed template
    # constants, not per-deal inputs — intentionally not mapped here.
    'refi_term_months': 69,
    'refi_interest_rate': 55,
    'refi_financing_costs_pct': 70,
    'refi_io_periods': 71,
    # Growth rates
    'rent_growth_rate': 56,
    'opex_growth_rate': 57,
    'exit_cap_rate': 59,
    # Waterfall
    'am_fee_pct': 64,
    'pari_passu': 90,
    'preferred_return_pct': 65,
    'gp_promote_pct': 66,
    # Property tax
    'assessed_value': 49,
    'assessed_value_next_buyer': 52,
    'assessment_pct': 51,
    'millage_rate': 91,
    'special_assessments': 92,
}


# ─── helpers ────────────────────────────────────────────────────────────────

def _to_decimal(val, fallback=0.0):
    """Normalize a %-or-decimal value to 0–1 form.
    Values > 1 are divided by 100 (e.g. 6.5 → 0.065).
    Values ≤ 1 are used as-is. None returns fallback.
    Not used for pure ratios like DSCR."""
    if val is None:
        return fallback
    try:
        f = float(val)
        return f / 100.0 if f > 1.0 else f
    except (TypeError, ValueError):
        return fallback


def _as_list(val):
    """Wrap a scalar in a list; pass lists through unchanged."""
    return val if isinstance(val, list) else [val]


def _write_5yr_rows(ws, rows, value_or_list):
    """Write a scalar or up to-5-element list into the given 5 Inputs rows (col C), one per year.
    Shorter lists are padded by repeating the last element."""
    if isinstance(value_or_list, list) and len(value_or_list) > 0:
        base = list(value_or_list)
        vals = (base + [base[-1]] * 5)[:5]
    else:
        v = value_or_list[0] if isinstance(value_or_list, list) else value_or_list
        vals = [v] * 5
    for row, v in zip(rows, vals):
        ws.cell(row=row, column=3).value = v


def _build_claude_content(file, prompt_text):
    """Build Claude message content for a PDF or Excel/CSV file."""
    filename_lower = file.filename.lower()
    if filename_lower.endswith('.pdf'):
        file_bytes = file.read()
        pdf_b64 = base64.b64encode(file_bytes).decode('utf-8')
        return [
            {
                'type': 'document',
                'source': {'type': 'base64', 'media_type': 'application/pdf', 'data': pdf_b64},
            },
            {
                'type': 'text',
                'text': prompt_text,
                'cache_control': {'type': 'ephemeral'},
            },
        ]
    else:
        import openpyxl
        file_bytes = file.read()
        text_parts = []
        if filename_lower.endswith('.csv'):
            import csv
            reader = csv.reader(io.StringIO(file_bytes.decode('utf-8', errors='replace')))
            for row in reader:
                line = '\t'.join(str(c) for c in row)
                if line.strip():
                    text_parts.append(line)
        else:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                text_parts.append(f'=== Sheet: {sheet_name} ===')
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else '' for c in row]
                    if any(c.strip() for c in cells):
                        text_parts.append('\t'.join(cells))
        doc_text = '\n'.join(text_parts)[:30000]
        return [{'type': 'text', 'text': f'Document content:\n\n{doc_text}\n\n---\n\n{prompt_text}'}]


def _call_claude(content, api_key):
    """Send content to Claude and return the parsed JSON dict."""
    from anthropic import Anthropic
    client = Anthropic(api_key=api_key)
    message = client.messages.create(
        model='claude-sonnet-4-6',
        max_tokens=4096,
        messages=[{'role': 'user', 'content': content}],
    )
    response_text = ''
    for block in message.content:
        if hasattr(block, 'text') and block.text:
            response_text = block.text.strip()
            break
    if not response_text:
        raise ValueError('Claude returned an empty response')
    if '```' in response_text:
        response_text = re.sub(r'^```json\s*', '', response_text, flags=re.MULTILINE)
        response_text = re.sub(r'^```\s*', '', response_text, flags=re.MULTILINE)
        response_text = re.sub(r'```$', '', response_text, flags=re.MULTILINE)
        response_text = response_text.strip()
    json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
    if not json_match:
        raise ValueError('No JSON object found in Claude response')
    return json.loads(json_match.group(0))


def _handle_extraction_error(e, route_name):
    msg = str(e)
    logger.error('[%s] error: %s', route_name, msg)
    if 'credit balance is too low' in msg or 'billing' in msg.lower():
        return jsonify({'success': False, 'error': 'Anthropic API credits exhausted.', 'code': 'BILLING_ERROR'}), 503
    if 'rate_limit' in msg.lower() or 'rate limit' in msg.lower():
        return jsonify({'success': False, 'error': 'Anthropic API rate limit. Try again shortly.', 'code': 'RATE_LIMIT'}), 429
    if 'invalid_api_key' in msg or 'authentication' in msg.lower():
        return jsonify({'success': False, 'error': 'Anthropic API key is invalid.', 'code': 'AUTH_ERROR'}), 503
    return jsonify({'success': False, 'error': msg, 'code': 'SERVER_ERROR'}), 500


# ─── /export ────────────────────────────────────────────────────────────────

@underwriting_v2_bp.route('/export', methods=['POST'])
def export():
    """Fill the v15 template's Inputs tab with scalar inputs and return the xlsx file.
    Python writes ONLY the Inputs tab — the Pro Forma tab is 100% formulas that read
    from Inputs, so Excel evaluates the whole model on open."""
    import openpyxl

    if not os.path.exists(TEMPLATE_PATH):
        return jsonify({'success': False, 'error': f'Template not found at {TEMPLATE_PATH}'}), 500

    body = request.get_json(silent=True) or {}

    wb = openpyxl.load_workbook(TEMPLATE_PATH, keep_links=False)
    ws = wb['Inputs']

    def set_row(row, value):
        ws.cell(row=row, column=3).value = value

    # ── Property ────────────────────────────────────────────────────
    if body.get('propertyName'):
        set_row(ROWS['property_name'], str(body['propertyName']))
    if body.get('address'):
        set_row(ROWS['address'], str(body['address']))

    # ── Timeline ────────────────────────────────────────────────────
    if body.get('acquisitionDate'):
        try:
            acq = datetime.strptime(str(body['acquisitionDate'])[:10], '%Y-%m-%d').date()
            set_row(ROWS['acquisition_date'], acq)
        except (ValueError, TypeError):
            pass

    bridge = body.get('bridgePeriodMonths') or body.get('bridgeLoanTermMonths')
    if bridge is not None:
        set_row(ROWS['bridge_term_months'], int(bridge))

    hold_yrs = body.get('holdPeriodYears') or body.get('holdingPeriod')
    if hold_yrs is not None:
        set_row(ROWS['hold_period_years'], int(hold_yrs))

    # ── Valuation ───────────────────────────────────────────────────
    ttm_noi = float(body['ttmNoi']) if body.get('ttmNoi') is not None else None
    purchase_price = float(body['purchasePrice']) if body.get('purchasePrice') is not None else None

    if ttm_noi is not None:
        set_row(ROWS['ttm_noi'], ttm_noi)
    if purchase_price is not None:
        set_row(ROWS['purchase_price'], purchase_price)

    # Cap Rate Aequitas Valuation — feeds the equity-basis formula on Pro Forma.
    # Always derive from actual NOI / purchase price so the model's implied basis matches
    # the confirmed purchase price. Fall back to the user-entered entry cap rate.
    if ttm_noi and purchase_price:
        set_row(ROWS['cap_rate_valuation'], ttm_noi / purchase_price)
    elif body.get('entryCapRate') is not None:
        set_row(ROWS['cap_rate_valuation'], _to_decimal(body['entryCapRate']))

    if body.get('salesCostPct') is not None:
        set_row(ROWS['sales_cost_pct'], _to_decimal(body['salesCostPct']))

    if body.get('workingCapitalReserveAmount') is not None:
        set_row(ROWS['working_capital_reserve'], float(body['workingCapitalReserveAmount']))

    # ── LP / GP split ───────────────────────────────────────────────
    if body.get('lpEquityShare') is not None:
        set_row(ROWS['lp_equity_share'], _to_decimal(body['lpEquityShare']))

    # ── Unit mix (cols B/C/D/E = type/count/rent/avgSf) ──────────────
    unit_mix = body.get('unitMix') or []
    for i in range(ROWS['unit_mix_rows']):
        row = ROWS['unit_mix_start'] + i
        if i < len(unit_mix):
            u = unit_mix[i]
            ws.cell(row=row, column=2).value = u.get('unitType') or u.get('type')
            ws.cell(row=row, column=3).value = int(u.get('count') or 0)
            ws.cell(row=row, column=4).value = float(u.get('askingRent') or u.get('rent') or 0)
            ws.cell(row=row, column=5).value = float(u.get('avgSf') or u.get('sqft') or 0)
        else:
            ws.cell(row=row, column=2).value = None
            ws.cell(row=row, column=3).value = 0
            ws.cell(row=row, column=4).value = 0
            ws.cell(row=row, column=5).value = 0

    # ── Other income ────────────────────────────────────────────────
    if body.get('rubsPct') is not None:
        set_row(ROWS['rubs_pct'], _to_decimal(body['rubsPct']))
    if body.get('parkingPerUnitMo') is not None:
        set_row(ROWS['parking_per_unit'], float(body['parkingPerUnitMo']))
    if body.get('otherIncomePerUnitMo') is not None:
        set_row(ROWS['other_income_per_unit'], float(body['otherIncomePerUnitMo']))

    # ── Income adjustments (5-year, one Inputs row per year) ────────
    def write_5yr_decimal(rows, val):
        _write_5yr_rows(ws, rows, [_to_decimal(v) for v in _as_list(val)])

    if body.get('lossToLeaseRate') is not None:
        write_5yr_decimal(ROWS['loss_to_lease_rows'], body['lossToLeaseRate'])
    if body.get('vacancyRate') is not None:
        write_5yr_decimal(ROWS['vacancy_rows'], body['vacancyRate'])
    if body.get('badDebtRate') is not None:
        write_5yr_decimal(ROWS['bad_debt_rows'], body['badDebtRate'])
    if body.get('concessionsRate') is not None:
        write_5yr_decimal(ROWS['concessions_rows'], body['concessionsRate'])
    if body.get('nonRevenueUnits') is not None:
        _write_5yr_rows(ws, ROWS['non_revenue_units_rows'], _as_list(body['nonRevenueUnits']))

    # ── Operating expenses ($/unit/yr) ───────────────────────────────
    opex_rows = {
        'opexPayrollPerUnit':         ROWS['opex_payroll_per_unit'],
        'opexAdminPerUnit':           ROWS['opex_admin_per_unit'],
        'opexMarketingPerUnit':       ROWS['opex_marketing_per_unit'],
        'opexRmPerUnit':              ROWS['opex_rm_per_unit'],
        'opexContractServicePerUnit': ROWS['opex_contract_service_per_unit'],
        'opexTurnoverPerUnit':        ROWS['opex_turnover_per_unit'],
        'opexOtherPerUnit':           ROWS['opex_other_per_unit'],
        'opexInsurancePerUnit':       ROWS['opex_insurance_per_unit'],
        'opexUtilitiesPerUnit':       ROWS['opex_utilities_per_unit'],
        'opexPropertyTaxPerUnit':     ROWS['opex_property_tax_per_unit'],
    }
    for field, row in opex_rows.items():
        if body.get(field) is not None:
            set_row(row, float(body[field]))

    # Management fee pct
    if body.get('managementFeePct') is not None:
        set_row(ROWS['management_fee_pct'], _to_decimal(body['managementFeePct']))

    # Cap reserve $/unit
    if body.get('capReservePerUnit') is not None:
        set_row(ROWS['cap_reserve_per_unit'], float(body['capReservePerUnit']))

    # ── Senior / bridge loan ─────────────────────────────────────────
    if body.get('seniorLoanAmount') is not None:
        set_row(ROWS['senior_loan_amount'], float(body['seniorLoanAmount']))
    if body.get('seniorInterestRate') is not None:
        set_row(ROWS['senior_interest_rate'], _to_decimal(body['seniorInterestRate']))
    if body.get('seniorFinancingCostsPct') is not None:
        set_row(ROWS['senior_financing_costs_pct'], _to_decimal(body['seniorFinancingCostsPct']))
    if body.get('seniorIoPeriods') is not None:
        set_row(ROWS['senior_io_periods'], int(body['seniorIoPeriods']))

    # ── Refinance loan ───────────────────────────────────────────────
    # Note: DSCR/DY/LTV sizing targets (F52/F53/F54 on Pro Forma) are fixed template
    # constants, not per-deal inputs — intentionally not written here.
    if body.get('refiTermMonths') is not None:
        set_row(ROWS['refi_term_months'], int(body['refiTermMonths']))
    if body.get('refiInterestRate') is not None:
        set_row(ROWS['refi_interest_rate'], _to_decimal(body['refiInterestRate']))
    if body.get('refiFinancingCostsPct') is not None:
        set_row(ROWS['refi_financing_costs_pct'], _to_decimal(body['refiFinancingCostsPct']))
    if body.get('refiIoPeriods') is not None:
        set_row(ROWS['refi_io_periods'], int(body['refiIoPeriods']))

    # ── Growth rates ──────────────────────────────────────────────────
    if body.get('rentGrowthRate') is not None:
        set_row(ROWS['rent_growth_rate'], _to_decimal(body['rentGrowthRate']))
    if body.get('opexGrowthRate') is not None:
        set_row(ROWS['opex_growth_rate'], _to_decimal(body['opexGrowthRate']))
    if body.get('exitCapRate') is not None:
        set_row(ROWS['exit_cap_rate'], _to_decimal(body['exitCapRate']))

    # ── Waterfall ────────────────────────────────────────────────────
    if body.get('amFeePct') is not None:
        set_row(ROWS['am_fee_pct'], _to_decimal(body['amFeePct']))
    if body.get('pariPassu') is not None:
        set_row(ROWS['pari_passu'], int(body['pariPassu']))
    if body.get('preferredReturnPct') is not None:
        set_row(ROWS['preferred_return_pct'], _to_decimal(body['preferredReturnPct']))
    if body.get('gpPromotePct') is not None:
        set_row(ROWS['gp_promote_pct'], _to_decimal(body['gpPromotePct']))

    # ── Property tax ─────────────────────────────────────────────────
    if body.get('assessedValue') is not None:
        set_row(ROWS['assessed_value'], float(body['assessedValue']))
    if body.get('assessedValueNextBuyer') is not None:
        set_row(ROWS['assessed_value_next_buyer'], float(body['assessedValueNextBuyer']))
    if body.get('assessmentPct') is not None:
        # 1.0 = 100% (full assessment) — do NOT divide by 100 when already ≤ 1
        set_row(ROWS['assessment_pct'], _to_decimal(body['assessmentPct']))
    if body.get('millageRate') is not None:
        set_row(ROWS['millage_rate'], _to_decimal(body['millageRate']))
    if body.get('specialAssessments') is not None:
        set_row(ROWS['special_assessments'], float(body['specialAssessments']))

    # ── Return file ──────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = re.sub(r'[^\w\-]', '_', body.get('propertyName') or 'deal')[:40]
    filename = f'Aequitas_Model_{safe_name}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


# ─── /extract-om ────────────────────────────────────────────────────────────

_OM_PROMPT = (
    'Respond with raw JSON only. No markdown, no backticks, no explanation, no preamble. '
    'Your entire response must be valid JSON starting with { and ending with }.\n\n'
    'This is a multifamily real estate Offering Memorandum (OM). '
    'Extract the following fields and return ONLY a valid JSON object.\n\n'
    'UNIT MIX RULES:\n'
    '- Count units from the individual rent roll rows, NOT from summary tables.\n'
    '- "count" = total units of that type including vacant rows.\n'
    '- "askingRent" = average monthly rent using ONLY occupied (non-vacant) rows for that type.\n'
    '- "avgSf" = average square footage for that unit type. REQUIRED — return 0 if not stated.\n'
    '- Use unitType labels like: "Studio", "1BR/1BA", "2BR/1BA", "2BR/2BA", "3BR/2BA".\n'
    '- Each entry in unitMix must represent ONE bedroom type.\n\n'
    'OPERATING EXPENSES RULES:\n'
    '- Pull annual totals from the operating expense summary table.\n'
    '- If only monthly figures are shown, multiply by 12 for annual.\n'
    '- managementFeePct: express as a decimal between 0 and 1 (e.g., 0.06 for 6%).\n\n'
    'VACANCY & CREDIT LOSS RULES:\n'
    '- vacancyRate: look for "vacancy", "vacancy loss", "physical vacancy". Express as decimal.\n'
    '- badDebtRate: look for "bad debt", "credit loss", "collection loss". Express as decimal.\n'
    '- lossToLeaseRate: gap between market rent and in-place rent as a % of GPR. Express as decimal.\n'
    '- concessionsRate: "concessions", "free rent", "lease-up concessions" as a % of GPR. Express as decimal.\n'
    '- If only a dollar amount is shown, divide by GPR to get the rate.\n\n'
    'OTHER INCOME RULES:\n'
    '- parkingIncomePerUnit: monthly parking income per unit in dollars.\n'
    '- rubsPct: RUBS or utility reimbursement percentage (0-100). Return as number 0-100, or null.\n'
    '- otherIncomePerUnit: monthly per-unit income excluding rent, parking, and laundry.\n'
    '- laundryIncome: total ANNUAL laundry and vending income only.\n\n'
    'PER-UNIT OPEX RULES:\n'
    '- Return annual dollars per unit for each line item.\n'
    '- opexPayrollPerUnit: payroll, salaries, benefits.\n'
    '- opexAdminPerUnit: administrative, office, legal, professional fees.\n'
    '- opexMarketingPerUnit: marketing, advertising, leasing.\n'
    '- opexRmPerUnit: repairs and maintenance only (not contract services).\n'
    '- opexContractServicePerUnit: contract services, landscaping, janitorial, pest control.\n'
    '- opexTurnoverPerUnit: turnover, make-ready costs.\n'
    '- capexPerUnit: capital expenditure reserve or replacement reserve.\n\n'
    'GROWTH RATES RULES:\n'
    '- opexGrowthRate: projected annual opex growth as a decimal. Null if not stated.\n'
    '- propertyTaxGrowthRate: projected annual property tax growth as a decimal. Null if not stated.\n\n'
    'PROPERTY TAX RULES:\n'
    '- assessedValue: the current assessed value of the property for tax purposes. Null if not found.\n'
    '- assessmentPct: the assessment ratio as a decimal (e.g., 1.0 = 100%, 0.85 = 85%). Null if not found.\n\n'
    'DEAL STRUCTURE RULES:\n'
    '- bridgePeriodMonths: renovation or value-add period length in months. Null if not stated.\n'
    '- lpEquityShare: LP equity share as a decimal (e.g., 0.85 for 85%). Null if not stated.\n'
    '- ttmNoi: trailing 12-month net operating income as a dollar figure. Look for "T-12 NOI", "TTM NOI", '
    '"trailing 12 NOI", "in-place NOI", or "current NOI". Null if not found.\n\n'
    'RENT STABILIZATION RULES:\n'
    '- rentStabilized: true if document mentions rent stabilization, rent control, or CPI caps.\n'
    '- annualRentGrowthCap: maximum allowed annual rent increase as a decimal.\n\n'
    'Return this exact JSON structure (use null for any field not found):\n'
    '{\n'
    '  "propertyName": "string or null",\n'
    '  "address": "full street address or null",\n'
    '  "city": "string or null",\n'
    '  "state": "2-letter state code or null",\n'
    '  "zipcode": "string or null",\n'
    '  "numUnits": "integer or null",\n'
    '  "askingPrice": "number or null",\n'
    '  "ttmNoi": "trailing 12-month NOI as number or null",\n'
    '  "unitMix": [\n'
    '    {"unitType": "e.g. 2BR/2BA", "count": "integer", "avgSf": "integer (0 if unknown)", "askingRent": "number or null"}\n'
    '  ],\n'
    '  "laundryIncome": "annual laundry/vending income as number or null",\n'
    '  "vacancyRate": "decimal 0.0-1.0 or null",\n'
    '  "badDebtRate": "decimal 0.0-1.0 or null",\n'
    '  "lossToLeaseRate": "decimal 0.0-1.0 or null",\n'
    '  "concessionsRate": "decimal 0.0-1.0 or null",\n'
    '  "parkingIncomePerUnit": "monthly $/unit or null",\n'
    '  "rubsPct": "RUBS % as number 0-100 or null",\n'
    '  "otherIncomePerUnit": "monthly $/unit (excl. parking and laundry) or null",\n'
    '  "operatingExpenses": {\n'
    '    "utilitiesAnnual": "number or null",\n'
    '    "insuranceAnnual": "number or null",\n'
    '    "propertyTaxAnnual": "number or null",\n'
    '    "repairsMaintenanceAnnual": "number or null",\n'
    '    "managementFeePct": "decimal 0.0-1.0 or null"\n'
    '  },\n'
    '  "opexPayrollPerUnit": "annual $/unit or null",\n'
    '  "opexAdminPerUnit": "annual $/unit or null",\n'
    '  "opexMarketingPerUnit": "annual $/unit or null",\n'
    '  "opexRmPerUnit": "annual $/unit or null",\n'
    '  "opexContractServicePerUnit": "annual $/unit or null",\n'
    '  "opexTurnoverPerUnit": "annual $/unit or null",\n'
    '  "capexPerUnit": "annual $/unit or null",\n'
    '  "opexGrowthRate": "annual opex growth as decimal or null",\n'
    '  "propertyTaxGrowthRate": "annual property tax growth as decimal or null",\n'
    '  "assessedValue": "number or null",\n'
    '  "assessmentPct": "decimal 0.0-1.0 or null",\n'
    '  "bridgePeriodMonths": "integer or null",\n'
    '  "lpEquityShare": "decimal 0.0-1.0 or null",\n'
    '  "rentStabilized": "boolean or null",\n'
    '  "annualRentGrowthCap": "decimal 0.0-1.0 or null"\n'
    '}'
)


@underwriting_v2_bp.route('/extract-om', methods=['POST'])
def extract_om():
    """Extract deal inputs from an Offering Memorandum PDF."""
    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        return jsonify({'success': False, 'error': 'ANTHROPIC_API_KEY not configured', 'code': 'AUTH_ERROR'}), 503

    file = request.files.get('file')
    if not file:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    try:
        content = _build_claude_content(file, _OM_PROMPT)
        data = _call_claude(content, api_key)
        logger.info('[extract-om v2] extracted: %s', json.dumps(data, indent=2))
        return jsonify({'success': True, 'data': data}), 200
    except Exception as e:
        return _handle_extraction_error(e, 'extract-om-v2')


# ─── /extract-rent-roll ──────────────────────────────────────────────────────

_RENT_ROLL_PROMPT = (
    'Respond with raw JSON only. No markdown, no backticks, no explanation, no preamble. '
    'Your entire response must be valid JSON starting with { and ending with }.\n\n'
    'This is a multifamily real estate Rent Roll document. '
    'Extract unit mix and rental data and return ONLY a valid JSON object.\n\n'
    'UNIT MIX RULES:\n'
    '- Count ALL units of each bedroom type, including vacant ones.\n'
    '- "count" = total units of that type (including vacant rows).\n'
    '- "askingRent" = average monthly rent using ONLY occupied (non-vacant) rows for that type.\n'
    '- "avgSf" = average square footage for that unit type. REQUIRED — return 0 if the rent roll '
    'does not include square footage.\n'
    '- Use unitType labels like: "Studio", "1BR/1BA", "2BR/1BA", "2BR/2BA", "3BR/2BA".\n'
    '- Each entry in unitMix must represent ONE bedroom type.\n\n'
    'VACANCY RULES:\n'
    '- vacancyRate: total vacant units / total units as a decimal (e.g. 0.05 for 5%).\n'
    '- badDebtRate: look for "bad debt", "credit loss", "collection loss" as a decimal.\n\n'
    'RENT STABILIZATION RULES:\n'
    '- rentStabilized: true if document mentions rent stabilization, rent control, or CPI caps.\n'
    '- annualRentGrowthCap: maximum allowed annual rent increase as a decimal.\n\n'
    'Return this exact JSON structure (use null for any field not found):\n'
    '{\n'
    '  "propertyName": "string or null",\n'
    '  "address": "full street address or null",\n'
    '  "city": "string or null",\n'
    '  "state": "2-letter state code or null",\n'
    '  "zipcode": "string or null",\n'
    '  "numUnits": "integer or null",\n'
    '  "askingPrice": null,\n'
    '  "unitMix": [\n'
    '    {"unitType": "e.g. 2BR/2BA", "count": "integer", "avgSf": "integer (0 if unknown)", "askingRent": "number or null"}\n'
    '  ],\n'
    '  "laundryIncome": null,\n'
    '  "vacancyRate": "decimal 0.0-1.0 or null",\n'
    '  "badDebtRate": "decimal 0.0-1.0 or null",\n'
    '  "operatingExpenses": null,\n'
    '  "rentStabilized": "boolean or null",\n'
    '  "annualRentGrowthCap": "decimal 0.0-1.0 or null"\n'
    '}'
)


@underwriting_v2_bp.route('/extract-rent-roll', methods=['POST'])
def extract_rent_roll():
    """Extract unit mix and vacancy data from a Rent Roll (PDF, xlsx, or csv)."""
    api_key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not api_key:
        return jsonify({'success': False, 'error': 'ANTHROPIC_API_KEY not configured', 'code': 'AUTH_ERROR'}), 503

    file = request.files.get('file')
    if not file:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    try:
        content = _build_claude_content(file, _RENT_ROLL_PROMPT)
        data = _call_claude(content, api_key)
        logger.info('[extract-rent-roll v2] extracted: %s', json.dumps(data, indent=2))
        return jsonify({'success': True, 'data': data}), 200
    except Exception as e:
        return _handle_extraction_error(e, 'extract-rent-roll-v2')
