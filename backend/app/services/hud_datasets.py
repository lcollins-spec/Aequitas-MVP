"""
HUD national bulk-file connector for the sourcing signals engine.

Unlike the county/city connectors in signal_connectors.py, these datasets
are national, single-file, and free — no per-market config beyond the
market's city/state is needed at all, which makes this the most literally
market-agnostic connector of the set.

Source URLs and column names below were verified by directly downloading and
inspecting each file (not guessed from documentation):

  - FHA-insured multifamily mortgages: a single flat file. Row 0 is a title
    row ("FHA_BF90_RM_A", count), the real header is row 1. Has PROPERTY
    NAME/CITY/STATE/ZIP and UNITS/MATURITY DATE — no street address column,
    so 'address' for this signal is a name+city+state string, not a street
    address.
  - Section 8 assistance: HUD publishes property info and contract info as
    TWO separate files joined by `property_id` — the property file has no
    expiration date, and the contract file has no address. Both are joined
    here to produce one hit with both address and expiration date.

Memory note: these files are loaded with read_only=True and a column
allowlist, not a full in-memory load. A full load of the largest file (23.6k
rows x 74 cols) measured at ~750MB peak RSS during a Render free-tier (512MB)
deploy that OOM'd — read_only mode plus column projection measured at ~90MB
for that same file. The FHA file's own worksheet metadata falsely declares a
1x1 dimension (a stale <dimension> tag), which read_only mode trusts to
decide when to stop streaming — iter_rows() is called with an explicit
oversized max_row/max_col below specifically to bypass that and read the
real ~15k rows regardless.

The LIHTC database is also a real, free, address-level public dataset.
HUD's own huduser.gov/lihtc/ page is an interactive query tool with no
static bulk file, but HUD separately publishes the same data as a plain
ArcGIS FeatureServer layer (HUD_LIHTC_ARCGIS_URL below) — found by searching
HUD's ArcGIS org rather than the huduser.gov page, verified with live
queries (119 real Sacramento projects, 26 real Columbus projects). Unlike
the three bulk-file datasets above, this one is queried per-market with a
server-side WHERE clause (city/state uppercased — the layer stores them
that way) rather than downloaded once and filtered client-side, since
Esri's query endpoint does the filtering for us. Watch for YR_PIS == 8888,
a placeholder the dataset uses for "unknown," not a real year.
"""
import io
import logging
from datetime import datetime, timedelta

import requests
import openpyxl

from app.services import signal_connectors

logger = logging.getLogger(__name__)

REQUEST_TIMEOUT = 60

HUD_FHA_MORTGAGES_URL = 'https://www.hud.gov/sites/default/files/Housing/documents/FHA-BF90-RM-A.xlsx'
HUD_SECTION8_PROPERTIES_URL = 'https://www.hud.gov/sites/dfiles/Housing/documents/MF-Properties-with-Assistance-Sec8-Contracts1.xlsx'
HUD_SECTION8_CONTRACTS_URL = 'https://www.hud.gov/sites/dfiles/Housing/documents/MF-Assistance-Sec8-Contracts1.xlsx'
HUD_LIHTC_ARCGIS_URL = 'https://egis.hud.gov/arcgis/rest/services/affht/AffhtMapService/MapServer/30'
LIHTC_YEAR15_UNKNOWN_YEAR = 8888

UNIT_MIN = 20
UNIT_MAX = 80

# Column allowlists — only these survive parsing, the rest are dropped
# immediately per-row rather than loaded and discarded later.
FHA_COLUMNS = [
    'PROPERTY NAME', 'PROPERTY CITY', 'PROPERTY STATE', 'PROPERTY ZIP',
    'UNITS', 'MATURITY DATE', 'ORIGINAL MORTGAGE AMOUNT', 'HOLDER NAME',
]
SECTION8_PROPERTY_COLUMNS = [
    'property_id', 'property_name_text', 'address_line1_text', 'city_name_text',
    'state_code', 'zip_code', 'property_total_unit_count',
    'owner_organization_name', 'owner_individual_full_name',
    'owner_address_line1', 'owner_city_name', 'owner_state_code', 'owner_zip_code',
]
SECTION8_CONTRACT_COLUMNS = [
    'property_id', 'contract_number', 'tracs_current_expiration_date',
    'tracs_overall_expiration_date', 'assisted_units_count', 'program_type_name',
]

# In-memory cache: {dataset_key: {'rows': [...], 'fetched_at': datetime}}
_cache = {}


def _clean(value):
    return value.strip() if isinstance(value, str) else value


def _download_excel_rows(url, header_row_index=0, keep_columns=None):
    resp = requests.get(url, timeout=REQUEST_TIMEOUT)
    resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True, read_only=True)
    ws = wb.active
    # max_row/max_col explicitly oversized: some HUD files declare a wrong
    # worksheet dimension that read_only mode otherwise trusts (see module
    # docstring) — this forces it to keep streaming to the real end of data.
    rows_iter = ws.iter_rows(min_row=1, max_row=10_000_000, max_col=500, values_only=True)
    for _ in range(header_row_index):
        next(rows_iter)
    headers = [str(h).strip() if h is not None else '' for h in next(rows_iter)]

    keep_idx = None
    if keep_columns is not None:
        keep_set = set(keep_columns)
        keep_idx = [i for i, h in enumerate(headers) if h in keep_set]

    rows = []
    for row in rows_iter:
        if keep_idx is not None:
            rows.append({headers[i]: _clean(row[i]) for i in keep_idx if i < len(row)})
        else:
            rows.append(dict(zip(headers, (_clean(v) for v in row))))
    wb.close()
    return rows


def refresh_hud_datasets(force=False):
    """Download and cache the national HUD files. Cheap no-op if already cached unless force=True."""
    if 'fha' not in _cache or force:
        try:
            rows = _download_excel_rows(HUD_FHA_MORTGAGES_URL, header_row_index=1, keep_columns=FHA_COLUMNS)
            _cache['fha'] = {'rows': rows, 'fetched_at': datetime.utcnow()}
            logger.info(f"Cached {len(rows)} rows from HUD FHA multifamily mortgages file")
        except Exception as e:
            logger.warning(f"HUD FHA dataset refresh failed: {e}")

    if 'section8_properties' not in _cache or force:
        try:
            rows = _download_excel_rows(HUD_SECTION8_PROPERTIES_URL, keep_columns=SECTION8_PROPERTY_COLUMNS)
            _cache['section8_properties'] = {'rows': rows, 'fetched_at': datetime.utcnow()}
            logger.info(f"Cached {len(rows)} rows from HUD Section 8 properties file")
        except Exception as e:
            logger.warning(f"HUD Section 8 properties dataset refresh failed: {e}")

    if 'section8_contracts' not in _cache or force:
        try:
            rows = _download_excel_rows(HUD_SECTION8_CONTRACTS_URL, keep_columns=SECTION8_CONTRACT_COLUMNS)
            _cache['section8_contracts'] = {'rows': rows, 'fetched_at': datetime.utcnow()}
            logger.info(f"Cached {len(rows)} rows from HUD Section 8 contracts file")
        except Exception as e:
            logger.warning(f"HUD Section 8 contracts dataset refresh failed: {e}")

    # LIHTC intentionally skipped — HUD_LIHTC_URL unset, see module docstring.


def _in_unit_range(units):
    if units in (None, ''):
        return True  # unknown unit count — surface for manual review rather than drop
    try:
        return UNIT_MIN <= float(units) <= UNIT_MAX
    except (TypeError, ValueError):
        return True


def _raw(row):
    return {k: (v.isoformat() if isinstance(v, datetime) else v) for k, v in row.items()}


def _sql_quote(value):
    """Escape single quotes for an ArcGIS SQL WHERE clause (not parametrized like a DB driver)."""
    return (value or '').replace("'", "''")


def scan_lihtc_for_market(market, year15_horizon_years=2):
    """
    Query the HUD LIHTC ArcGIS layer directly for this market's city/state
    (server-side WHERE, no bulk download) and flag properties approaching or
    past Year 15 (when LIHTC compliance restrictions ease and owners commonly
    sell). Returns a list of normalized hit dicts (not yet persisted).
    """
    city = _sql_quote((market.city or '').strip().upper())
    state = _sql_quote((market.state or '').strip().upper())
    if not city or not state:
        return []

    where = f"PROJ_CTY='{city}' AND PROJ_ST='{state}'"
    field_mapping = {
        'project_name': 'PROJECT', 'address': 'PROJ_ADD', 'units': 'N_UNITS',
        'year_placed_in_service': 'YR_PIS', 'owner_name': 'COMPANY',
    }
    try:
        records = signal_connectors.fetch_arcgis_feature_server(HUD_LIHTC_ARCGIS_URL, field_mapping, where=where)
    except Exception as e:
        logger.warning(f"HUD LIHTC query failed for market {market.id}: {e}")
        return []

    current_year = datetime.utcnow().year
    hits = []
    for rec in records:
        units = rec.get('units')
        if not _in_unit_range(units):
            continue
        year_pis = rec.get('year_placed_in_service')
        if year_pis in (None, '', LIHTC_YEAR15_UNKNOWN_YEAR):
            continue  # unknown placed-in-service year — can't compute Year 15, skip rather than guess
        try:
            years_since = current_year - int(year_pis)
        except (TypeError, ValueError):
            continue
        if years_since < (15 - year15_horizon_years):
            continue  # not yet approaching Year 15
        address = rec.get('address') or rec.get('project_name') or ''
        owner_name = rec.get('owner_name')
        hits.append({
            'source': 'hud_lihtc_year15',
            'address': str(address).strip(),
            'owner_name': str(owner_name).strip() if owner_name and str(owner_name).strip() else None,
            'unit_count': int(float(units)) if units not in (None, '') else None,
            'raw_data': rec,
        })
    return hits


def scan_hud_signals_for_market(market, maturity_horizon_months=24):
    """
    Filter cached HUD rows down to this market's city/state + the 20-80 unit
    range, and flag properties nearing loan maturity or Section 8 contract
    expiration. Returns a list of normalized hit dicts (not yet persisted).
    """
    hits = []
    now = datetime.utcnow()
    horizon = now + timedelta(days=30 * maturity_horizon_months)
    city = (market.city or '').strip().lower()
    state = (market.state or '').strip().lower()

    fha = _cache.get('fha')
    if fha:
        for row in fha['rows']:
            if str(row.get('PROPERTY CITY') or '').strip().lower() != city:
                continue
            if str(row.get('PROPERTY STATE') or '').strip().lower() != state:
                continue
            units = row.get('UNITS')
            if not _in_unit_range(units):
                continue
            maturity = row.get('MATURITY DATE')
            if isinstance(maturity, datetime) and not (now <= maturity <= horizon):
                continue
            name = row.get('PROPERTY NAME') or ''
            address = ', '.join(filter(None, [str(name).strip(), row.get('PROPERTY CITY'), row.get('PROPERTY STATE')]))
            hits.append({
                'source': 'hud_fha_loan_maturity',
                'address': address,
                'unit_count': int(float(units)) if units not in (None, '') else None,
                'raw_data': _raw(row),
            })

    props = _cache.get('section8_properties')
    contracts = _cache.get('section8_contracts')
    if props and contracts:
        props_by_id = {}
        for row in props['rows']:
            if str(row.get('city_name_text') or '').strip().lower() != city:
                continue
            if str(row.get('state_code') or '').strip().lower() != state:
                continue
            pid = row.get('property_id')
            if pid is not None:
                props_by_id[pid] = row

        if props_by_id:
            for row in contracts['rows']:
                prop = props_by_id.get(row.get('property_id'))
                if not prop:
                    continue
                units = prop.get('property_total_unit_count')
                if not _in_unit_range(units):
                    continue
                expiration = row.get('tracs_current_expiration_date') or row.get('tracs_overall_expiration_date')
                if isinstance(expiration, datetime) and not (now <= expiration <= horizon):
                    continue
                address = ', '.join(filter(None, [
                    prop.get('address_line1_text'), prop.get('city_name_text'), prop.get('state_code'),
                ]))
                owner_name = prop.get('owner_organization_name') or prop.get('owner_individual_full_name')
                owner_mailing = ', '.join(filter(None, [
                    prop.get('owner_address_line1'), prop.get('owner_city_name'), prop.get('owner_state_code'),
                ]))
                hits.append({
                    'source': 'hud_section8_contract_expiration',
                    'address': address,
                    'owner_name': str(owner_name).strip() if owner_name else None,
                    'owner_mailing_address': owner_mailing or None,
                    'unit_count': int(float(units)) if units not in (None, '') else None,
                    'raw_data': {
                        **{f'contract_{k}': v for k, v in _raw(row).items()},
                        **{f'property_{k}': v for k, v in _raw(prop).items()},
                    },
                })

    return hits
