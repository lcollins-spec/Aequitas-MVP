#!/usr/bin/env python3
"""
Aequitas Multifamily Underwriting Model Builder
Builds a production-ready Excel underwriting model for NOAH acquisitions
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from datetime import datetime, timedelta
import math

# Global workbook reference for named ranges
_workbook = None

# Constants for styling
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
CALC_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
POSITIVE_FILL = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
NEGATIVE_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

HEADER_FONT = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
TITLE_FONT = Font(name='Calibri', size=18, bold=True)
BOLD_FONT = Font(name='Calibri', size=11, bold=True)
NORMAL_FONT = Font(name='Calibri', size=11)

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_underwriting_model(data=None):
    """
    Create the complete Aequitas underwriting Excel model

    Args:
        data: Optional dictionary containing multifamily underwriting data.
              If provided, will populate the model with real data.
              If None, will generate a template with sample data.
    """
    global _workbook

    wb = Workbook()
    _workbook = wb  # Set global reference
    wb.remove(wb.active)  # Remove default sheet

    # Create tabs in order
    ws_assumptions = wb.create_sheet("ASSUMPTIONS")
    ws_sources_uses = wb.create_sheet("SOURCES & USES")
    ws_debt = wb.create_sheet("DEBT SCHEDULE")
    ws_annual_cf = wb.create_sheet("ANNUAL CASH FLOW")
    ws_reference = wb.create_sheet("REFERENCE DATA")

    # Build each tab (pass data to populate with real values)
    build_assumptions_tab(ws_assumptions, data)
    build_sources_uses_tab(ws_sources_uses)
    build_debt_schedule_tab(ws_debt)
    build_annual_cashflow_tab(ws_annual_cf)
    build_reference_tab(ws_reference)

    # Hide reference tab
    ws_reference.sheet_state = 'hidden'

    return wb

def build_assumptions_tab(ws, data=None):
    """
    Build the ASSUMPTIONS tab with all user inputs

    Args:
        ws: The worksheet to build
        data: Optional dictionary with real property data
    """
    # Use provided data or defaults
    if data is None:
        data = {}

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    row = 1

    # Title
    ws[f'A{row}'] = "AEQUITAS BEDROCK HOUSING - UNDERWRITING ASSUMPTIONS"
    ws[f'A{row}'].font = TITLE_FONT
    ws.merge_cells(f'A{row}:D{row}')
    row += 2

    # === PROPERTY INFORMATION ===
    ws[f'A{row}'] = "PROPERTY INFORMATION"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    # Property details
    add_input_row(ws, row, "Property Name", data.get('propertyName', 'Sample Multifamily Property'), "B"); row += 1
    add_input_row(ws, row, "Address", data.get('address', '123 Main Street'), "B"); row += 1
    add_input_row(ws, row, "City", data.get('city', 'Sacramento'), "B"); row += 1
    add_input_row(ws, row, "County", data.get('county', 'Sacramento'), "B"); row += 1
    add_input_row(ws, row, "State", data.get('state', 'California'), "B"); row += 1
    add_input_row(ws, row, "ZIP Code", data.get('zipCode', '95814'), "B"); row += 1
    add_input_row(ws, row, "Year Built", data.get('yearBuilt', 1985), "B", num_format='0'); row += 1
    add_input_row(ws, row, "Building Type", data.get('buildingType', 'Garden Style'), "B"); row += 1
    add_input_row(ws, row, "Number of Buildings", data.get('numberOfBuildings', 4), "B", num_format='0'); row += 1
    add_input_row(ws, row, "Parking Spaces", data.get('parkingSpaces', 120), "B", num_format='0'); row += 1
    row += 1

    # === ACQUISITION ===
    ws[f'A{row}'] = "ACQUISITION"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    add_input_row(ws, row, "Purchase Price", data.get('purchasePrice', 10000000), "B", num_format='$#,##0', name="Purchase_Price"); row += 1

    # Parse acquisition date (remove timezone for Excel compatibility)
    acq_date = data.get('acquisitionDate', datetime(2025, 3, 1))
    if isinstance(acq_date, str):
        from datetime import datetime as dt
        acq_date = dt.fromisoformat(acq_date.replace('Z', '+00:00'))
        # Remove timezone info for Excel compatibility
        acq_date = acq_date.replace(tzinfo=None)
    add_input_row(ws, row, "Acquisition Date", acq_date, "B", num_format='mm/dd/yyyy'); row += 1

    add_input_row(ws, row, "Earnest Money %", data.get('earnestMoneyPct', 0.02), "B", num_format='0.0%'); row += 1
    add_input_row(ws, row, "Closing Costs %", data.get('closingCostsPct', 0.03), "B", num_format='0.0%'); row += 1
    add_input_row(ws, row, "Due Diligence Costs", data.get('dueDiligenceCosts', 50000), "B", num_format='$#,##0'); row += 1
    row += 1

    # === UNIT MIX ===
    ws[f'A{row}'] = "UNIT MIX"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    # Unit mix table header
    headers = ['Unit Type', 'Count', 'Avg SF', 'Current Rent', 'Market Rent', 'Reno Cost/Unit']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row, col_idx, header)
        cell.font = BOLD_FONT
        cell.fill = INPUT_FILL
        cell.border = THIN_BORDER
    row += 1

    # Unit mix data - use from data or defaults
    unit_mix_data = data.get('unitMix', [
        {'unitType': 'Studio', 'count': 10, 'avgSf': 500, 'currentRent': 1100, 'marketRent': 1300, 'renovationCostPerUnit': 6000},
        {'unitType': '1BR/1BA', 'count': 30, 'avgSf': 700, 'currentRent': 1400, 'marketRent': 1700, 'renovationCostPerUnit': 7000},
        {'unitType': '2BR/1BA', 'count': 25, 'avgSf': 900, 'currentRent': 1700, 'marketRent': 2100, 'renovationCostPerUnit': 8000},
        {'unitType': '2BR/2BA', 'count': 25, 'avgSf': 1000, 'currentRent': 1900, 'marketRent': 2400, 'renovationCostPerUnit': 9000},
        {'unitType': '3BR/2BA', 'count': 10, 'avgSf': 1200, 'currentRent': 2200, 'marketRent': 2700, 'renovationCostPerUnit': 10000}
    ])

    # Convert to list of tuples if it's a list of dicts
    unit_types = []
    for unit in unit_mix_data:
        if isinstance(unit, dict):
            unit_types.append((
                unit.get('unitType', ''),
                unit.get('count', 0),
                unit.get('avgSf', 0),
                unit.get('currentRent', 0),
                unit.get('marketRent', 0),
                unit.get('renovationCostPerUnit', 0)
            ))
        else:
            unit_types.append(unit)

    unit_mix_start_row = row
    for unit_data in unit_types:
        for col_idx, value in enumerate(unit_data, start=1):
            cell = ws.cell(row, col_idx, value)
            if col_idx == 1:  # Unit type
                cell.fill = CALC_FILL
            else:  # Numbers
                cell.fill = INPUT_FILL
                if col_idx in [2, 3]:  # Count, Avg SF
                    cell.number_format = '#,##0'
                elif col_idx in [4, 5, 6]:  # Rents and reno cost
                    cell.number_format = '$#,##0'
            cell.border = THIN_BORDER
        row += 1

    unit_mix_end_row = row - 1

    # Totals row
    cell = ws.cell(row, 1, "TOTALS")
    cell.font = BOLD_FONT
    cell.border = THIN_BORDER

    # Total units
    cell = ws.cell(row, 2, f'=SUM(B{unit_mix_start_row}:B{unit_mix_end_row})')
    cell.font = BOLD_FONT
    cell.number_format = '#,##0'
    cell.border = THIN_BORDER
    add_named_range(ws, f'B${row}', "Total_Units")

    # Weighted avg SF
    cell = ws.cell(row, 3, f'=SUMPRODUCT(B{unit_mix_start_row}:B{unit_mix_end_row},C{unit_mix_start_row}:C{unit_mix_end_row})/B{row}')
    cell.font = BOLD_FONT
    cell.number_format = '#,##0'
    cell.border = THIN_BORDER

    # Weighted avg current rent
    cell = ws.cell(row, 4, f'=SUMPRODUCT(B{unit_mix_start_row}:B{unit_mix_end_row},D{unit_mix_start_row}:D{unit_mix_end_row})/B{row}')
    cell.font = BOLD_FONT
    cell.number_format = '$#,##0'
    cell.border = THIN_BORDER
    add_named_range(ws, f'D${row}', "Avg_Current_Rent")

    # Weighted avg market rent
    cell = ws.cell(row, 5, f'=SUMPRODUCT(B{unit_mix_start_row}:B{unit_mix_end_row},E{unit_mix_start_row}:E{unit_mix_end_row})/B{row}')
    cell.font = BOLD_FONT
    cell.number_format = '$#,##0'
    cell.border = THIN_BORDER
    add_named_range(ws, f'E${row}', "Avg_Market_Rent")

    # Total renovation cost
    cell = ws.cell(row, 6, f'=SUMPRODUCT(B{unit_mix_start_row}:B{unit_mix_end_row},F{unit_mix_start_row}:F{unit_mix_end_row})')
    cell.font = BOLD_FONT
    cell.number_format = '$#,##0'
    cell.border = THIN_BORDER
    add_named_range(ws, f'F${row}', "Total_Unit_Reno_Cost")
    row += 2

    # === CURRENT OPERATIONS (T12) ===
    ws[f'A{row}'] = "CURRENT OPERATIONS (T12)"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    add_input_row(ws, row, "Physical Occupancy %", data.get('physicalOccupancy', 0.90), "B", num_format='0.0%'); row += 1
    add_input_row(ws, row, "Economic Occupancy %", data.get('economicOccupancy', 0.87), "B", num_format='0.0%'); row += 1

    # Gross potential rent (calculated)
    ws[f'A{row}'] = "Gross Potential Rent (Annual)"
    ws[f'B{row}'] = '=Total_Units*Avg_Current_Rent*12'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].fill = CALC_FILL
    add_named_range(ws, f'B${row}', "T12_GPR")
    row += 1

    # Other income
    other_income = data.get('otherIncome', {})
    ws[f'A{row}'] = "Other Income (Annual):"
    ws[f'A{row}'].font = BOLD_FONT
    row += 1
    add_input_row(ws, row, "  Laundry ($/unit/month)", other_income.get('laundryPerUnit', 15), "B", num_format='$#,##0'); row += 1
    add_input_row(ws, row, "  Pet Rent ($/unit/month)", other_income.get('petRentPerUnit', 25), "B", num_format='$#,##0'); row += 1
    add_input_row(ws, row, "  Parking ($/space/month)", other_income.get('parkingPerSpace', 30), "B", num_format='$#,##0'); row += 1
    add_input_row(ws, row, "  Other ($/unit/month)", other_income.get('otherPerUnit', 10), "B", num_format='$#,##0'); row += 1

    parking_spaces = data.get('parkingSpaces', 120)
    ws[f'A{row}'] = "Total Other Income (Annual)"
    ws[f'B{row}'] = f'=(B{row-4}+B{row-3}+B{row-1})*Total_Units*12+B{row-2}*{parking_spaces}*12'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].fill = CALC_FILL
    ws[f'B{row}'].font = BOLD_FONT
    add_named_range(ws, f"B${row}", "T12_Other_Income")
    row += 1

    # Economic loss
    add_input_row(ws, row, "Vacancy Loss (Annual)", data.get('vacancyLossAnnual', 150000), "B", num_format='$#,##0'); row += 1
    add_input_row(ws, row, "Concessions (Annual)", data.get('concessionsAnnual', 20000), "B", num_format='$#,##0'); row += 1
    add_input_row(ws, row, "Bad Debt (Annual)", data.get('badDebtAnnual', 25000), "B", num_format='$#,##0'); row += 1

    # EGI
    ws[f'A{row}'] = "Effective Gross Income (EGI)"
    ws[f'B{row}'] = f'=T12_GPR+T12_Other_Income-B{row-3}-B{row-2}-B{row-1}'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].fill = CALC_FILL
    ws[f'B{row}'].font = BOLD_FONT
    add_named_range(ws, f"B${row}", "T12_EGI")
    row += 2

    # === OPERATING EXPENSES (T12) ===
    ws[f'A{row}'] = "OPERATING EXPENSES (T12)"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    opex_data = data.get('operatingExpenses', {})
    opex_start_row = row
    add_input_row(ws, row, "Property Taxes", opex_data.get('propertyTax', 110000), "B", num_format='$#,##0'); row += 1
    add_input_row(ws, row, "Insurance", opex_data.get('insurance', 55000), "B", num_format='$#,##0'); row += 1
    add_input_row(ws, row, "Utilities - Electric", opex_data.get('utilitiesElectric', 35000), "B", num_format='$#,##0'); row += 1
    add_input_row(ws, row, "Utilities - Gas", opex_data.get('utilitiesGas', 20000), "B", num_format='$#,##0'); row += 1
    add_input_row(ws, row, "Utilities - Water/Sewer", opex_data.get('utilitiesWaterSewer', 45000), "B", num_format='$#,##0'); row += 1
    add_input_row(ws, row, "Utilities - Trash", opex_data.get('utilitiesTrash', 15000), "B", num_format='$#,##0'); row += 1
    add_input_row(ws, row, "Repairs & Maintenance", opex_data.get('repairsMaintenance', 60000), "B", num_format='$#,##0'); row += 1
    add_input_row(ws, row, "Payroll", opex_data.get('payroll', 75000), "B", num_format='$#,##0'); row += 1
    add_input_row(ws, row, "Management Fee %", opex_data.get('managementFeePct', 0.04), "B", num_format='0.0%', name="Mgmt_Fee_Pct"); row += 1
    ws[f'A{row}'] = "  Management Fee ($)"
    ws[f'B{row}'] = '=T12_EGI*Mgmt_Fee_Pct'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].fill = CALC_FILL
    row += 1
    add_input_row(ws, row, "Marketing/Advertising", opex_data.get('marketing', 8000), "B", num_format='$#,##0'); row += 1
    add_input_row(ws, row, "Legal/Professional", opex_data.get('legalProfessional', 10000), "B", num_format='$#,##0'); row += 1
    add_input_row(ws, row, "Administrative", opex_data.get('administrative', 12000), "B", num_format='$#,##0'); row += 1
    opex_end_row = row

    ws[f'A{row+1}'] = "Total Operating Expenses"
    # Sum all opex except mgmt fee row (it's calculated)
    ws[f'B{row+1}'] = f'=B{opex_start_row}+B{opex_start_row+1}+B{opex_start_row+2}+B{opex_start_row+3}+B{opex_start_row+4}+B{opex_start_row+5}+B{opex_start_row+6}+B{opex_start_row+7}+B{opex_start_row+9}+B{opex_start_row+10}+B{opex_start_row+11}+B{opex_start_row+12}'
    ws[f'B{row+1}'].number_format = '$#,##0'
    ws[f'B{row+1}'].fill = CALC_FILL
    ws[f'B{row+1}'].font = BOLD_FONT
    ws[f'B{row+1}'].border = Border(top=Side(style='double'), bottom=Side(style='double'))
    add_named_range(ws, f'B${row+1}', "T12_OpEx")
    row += 2

    ws[f'A{row}'] = "Net Operating Income (NOI)"
    ws[f'B{row}'] = '=T12_EGI-T12_OpEx'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].fill = CALC_FILL
    ws[f'B{row}'].font = BOLD_FONT
    add_named_range(ws, f"B${row}", "T12_NOI")
    row += 1

    ws[f'A{row}'] = "Operating Expense Ratio"
    ws[f'B{row}'] = '=T12_OpEx/T12_EGI'
    ws[f'B{row}'].number_format = '0.0%'
    ws[f'B{row}'].fill = CALC_FILL
    row += 2

    # === RENOVATION BUDGET ===
    ws[f'A{row}'] = "RENOVATION BUDGET"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    ws[f'A{row}'] = "Per-Unit Interior (from Unit Mix)"
    ws[f'B{row}'] = '=Total_Unit_Reno_Cost'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].fill = CALC_FILL
    row += 1

    reno_budget = data.get('renovationBudget', {})
    add_input_row(ws, row, "Common Area/Exterior", reno_budget.get('commonAreaExterior', 100000), "B", num_format='$#,##0', name="Common_Area_Reno"); row += 1
    add_input_row(ws, row, "Contingency %", reno_budget.get('contingencyPct', 0.10), "B", num_format='0.0%', name="Reno_Contingency_Pct"); row += 1

    ws[f'A{row}'] = "Total Renovation Budget"
    ws[f'B{row}'] = '=(Total_Unit_Reno_Cost+Common_Area_Reno)*(1+Reno_Contingency_Pct)'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].fill = CALC_FILL
    ws[f'B{row}'].font = BOLD_FONT
    ws[f'B{row}'].border = Border(top=Side(style='double'), bottom=Side(style='double'))
    add_named_range(ws, f"B${row}", "Total_Renovation_Budget")
    row += 2

    # === OPERATING PROJECTIONS ===
    ws[f'A{row}'] = "OPERATING PROJECTIONS"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    op_proj = data.get('operatingProjections', {})
    add_input_row(ws, row, "Market Rent Growth % (Annual)", op_proj.get('marketRentGrowth', 0.03), "B", num_format='0.0%', name="Market_Rent_Growth"); row += 1
    add_input_row(ws, row, "In-Place Rent Growth % (Annual)", op_proj.get('inplaceRentGrowth', 0.025), "B", num_format='0.0%', name="InPlace_Rent_Growth"); row += 1
    add_input_row(ws, row, "Other Income Growth % (Annual)", op_proj.get('otherIncomeGrowth', 0.03), "B", num_format='0.0%', name="Other_Income_Growth"); row += 1
    add_input_row(ws, row, "OpEx Growth % (Annual)", op_proj.get('opexGrowth', 0.03), "B", num_format='0.0%', name="OpEx_Growth"); row += 1
    add_input_row(ws, row, "Stabilized Vacancy %", op_proj.get('stabilizedVacancy', 0.05), "B", num_format='0.0%', name="Stabilized_Vacancy"); row += 1
    add_input_row(ws, row, "CapEx Reserve ($/unit/year)", op_proj.get('capexPerUnitAnnual', 400), "B", num_format='$#,##0', name="CapEx_Per_Unit"); row += 1
    row += 1

    # === FINANCING ===
    ws[f'A{row}'] = "FINANCING"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    financing = data.get('financing', {})
    add_input_row(ws, row, "Loan Type", financing.get('loanType', 'Agency Fixed'), "B"); row += 1
    add_input_row(ws, row, "LTV %", financing.get('ltv', 0.70), "B", num_format='0.0%', name="LTV"); row += 1
    add_input_row(ws, row, "Interest Rate %", financing.get('interestRate', 0.06), "B", num_format='0.00%', name="Interest_Rate"); row += 1
    add_input_row(ws, row, "Amortization (years)", financing.get('amortizationYears', 30), "B", num_format='0', name="Amortization_Years"); row += 1
    add_input_row(ws, row, "Loan Term (years)", financing.get('loanTermYears', 10), "B", num_format='0'); row += 1
    add_input_row(ws, row, "Origination Fee %", financing.get('originationFeePct', 0.01), "B", num_format='0.0%'); row += 1
    add_input_row(ws, row, "Lender Legal/DD", financing.get('lenderLegalDd', 25000), "B", num_format='$#,##0'); row += 1
    row += 1

    ws[f'A{row}'] = "Loan Amount"
    ws[f'B{row}'] = '=Purchase_Price*LTV'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].fill = CALC_FILL
    ws[f'B{row}'].font = BOLD_FONT
    add_named_range(ws, f"B${row}", "Loan_Amount")
    row += 2

    # === EXIT ASSUMPTIONS ===
    ws[f'A{row}'] = "EXIT ASSUMPTIONS"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    exit_assump = data.get('exitAssumptions', {})
    add_input_row(ws, row, "Hold Period (years)", exit_assump.get('holdPeriodYears', 5), "B", num_format='0', name="Hold_Period"); row += 1
    add_input_row(ws, row, "Exit Cap Rate %", exit_assump.get('exitCapRate', 0.055), "B", num_format='0.00%', name="Exit_Cap_Rate"); row += 1
    add_input_row(ws, row, "Sale Costs %", exit_assump.get('saleCostsPct', 0.04), "B", num_format='0.0%', name="Sale_Costs_Pct"); row += 1
    row += 1

    # === PROPERTY TAX ===
    ws[f'A{row}'] = "PROPERTY TAX"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    prop_tax = data.get('propertyTax', {})
    add_input_row(ws, row, "County Tax Rate %", prop_tax.get('countyTaxRate', 0.011), "B", num_format='0.00%', name="Tax_Rate"); row += 1
    add_input_row(ws, row, "Prop 13 Cap % (CA only)", prop_tax.get('prop13Cap', 0.02), "B", num_format='0.0%', name="Prop13_Cap"); row += 1
    add_input_row(ws, row, "Mello-Roos/Special Assessments", prop_tax.get('specialAssessments', 0), "B", num_format='$#,##0', name="Special_Assessments"); row += 1

def add_named_range(ws, cell_address, name):
    """Helper to add a named range to a cell"""
    global _workbook
    if name and _workbook:
        from openpyxl.workbook.defined_name import DefinedName
        defn = DefinedName(name, attr_text=f"'{ws.title}'!${cell_address}")
        _workbook.defined_names.add(defn)

def add_input_row(ws, row, label, value, value_col, num_format=None, name=None):
    """Helper to add an input row with label and value"""
    ws[f'A{row}'] = label
    ws[f'{value_col}{row}'] = value
    ws[f'{value_col}{row}'].fill = INPUT_FILL
    ws[f'{value_col}{row}'].border = THIN_BORDER
    if num_format:
        ws[f'{value_col}{row}'].number_format = num_format
    if name:
        add_named_range(ws, f'{value_col}${row}', name)

def build_sources_uses_tab(ws):
    """Build the SOURCES & USES tab"""

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15

    row = 1

    # Title
    ws[f'A{row}'] = "SOURCES & USES OF FUNDS"
    ws[f'A{row}'].font = TITLE_FONT
    ws.merge_cells(f'A{row}:C{row}')
    row += 2

    # === SOURCES ===
    ws[f'A{row}'] = "SOURCES"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL
    ws.merge_cells(f'A{row}:C{row}')
    row += 1

    ws[f'A{row}'] = "Loan Proceeds"
    ws[f'B{row}'] = '=Loan_Amount'
    ws[f'B{row}'].number_format = '$#,##0'
    row += 1

    ws[f'A{row}'] = "Equity"
    equity_row = row
    ws[f'B{row}'] = f'=B{row+4}-B{row-1}'  # Total Uses - Loan = Equity
    ws[f'B{row}'].number_format = '$#,##0'
    add_named_range(ws, f"B${row}", "Total_Equity")
    row += 1

    ws[f'A{row}'] = "TOTAL SOURCES"
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'B{row}'] = f'=B{equity_row-1}+B{equity_row}'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].font = BOLD_FONT
    ws[f'B{row}'].border = Border(top=Side(style='double'), bottom=Side(style='double'))
    row += 2

    # === USES ===
    ws[f'A{row}'] = "USES"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL
    ws.merge_cells(f'A{row}:C{row}')
    row += 1

    # Acquisition costs
    ws[f'A{row}'] = "Acquisition Costs"
    ws[f'A{row}'].font = BOLD_FONT
    row += 1

    uses_start = row
    ws[f'A{row}'] = "  Purchase Price"
    ws[f'B{row}'] = '=Purchase_Price'
    ws[f'B{row}'].number_format = '$#,##0'
    row += 1

    ws[f'A{row}'] = "  Closing Costs"
    ws[f'B{row}'] = '=Purchase_Price*ASSUMPTIONS!B18'  # Closing costs %
    ws[f'B{row}'].number_format = '$#,##0'
    row += 1

    ws[f'A{row}'] = "  Due Diligence"
    ws[f'B{row}'] = '=ASSUMPTIONS!B19'
    ws[f'B{row}'].number_format = '$#,##0'
    row += 1

    ws[f'A{row}'] = "Subtotal Acquisition"
    ws[f'B{row}'] = f'=SUM(B{uses_start}:B{row-1})'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].font = BOLD_FONT
    subtotal_acq_row = row
    row += 2

    # Renovation costs
    ws[f'A{row}'] = "Renovation Costs"
    ws[f'A{row}'].font = BOLD_FONT
    row += 1

    ws[f'A{row}'] = "  Total Renovation Budget"
    ws[f'B{row}'] = '=Total_Renovation_Budget'
    ws[f'B{row}'].number_format = '$#,##0'
    subtotal_reno_row = row
    row += 2

    # Financing costs
    ws[f'A{row}'] = "Financing Costs"
    ws[f'A{row}'].font = BOLD_FONT
    row += 1

    fin_start = row
    ws[f'A{row}'] = "  Loan Origination Fee"
    ws[f'B{row}'] = '=Loan_Amount*ASSUMPTIONS!B95'  # Origination fee %
    ws[f'B{row}'].number_format = '$#,##0'
    row += 1

    ws[f'A{row}'] = "  Lender Legal/DD"
    ws[f'B{row}'] = '=ASSUMPTIONS!B96'
    ws[f'B{row}'].number_format = '$#,##0'
    row += 1

    ws[f'A{row}'] = "Subtotal Financing"
    ws[f'B{row}'] = f'=SUM(B{fin_start}:B{row-1})'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].font = BOLD_FONT
    subtotal_fin_row = row
    row += 2

    # Total uses
    ws[f'A{row}'] = "TOTAL USES"
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'B{row}'] = f'=B{subtotal_acq_row}+B{subtotal_reno_row}+B{subtotal_fin_row}'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].font = BOLD_FONT
    ws[f'B{row}'].border = Border(top=Side(style='double'), bottom=Side(style='double'))
    add_named_range(ws, f"B${row}", "Total_Uses")
    row += 2

    # Per unit metrics
    ws[f'A{row}'] = "PER UNIT METRICS"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL
    ws.merge_cells(f'A{row}:C{row}')
    row += 1

    ws[f'A{row}'] = "Purchase Price per Unit"
    ws[f'B{row}'] = '=Purchase_Price/Total_Units'
    ws[f'B{row}'].number_format = '$#,##0'
    row += 1

    ws[f'A{row}'] = "Renovation Cost per Unit"
    ws[f'B{row}'] = '=Total_Renovation_Budget/Total_Units'
    ws[f'B{row}'].number_format = '$#,##0'
    row += 1

    ws[f'A{row}'] = "Total Project Cost per Unit"
    ws[f'B{row}'] = '=Total_Uses/Total_Units'
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].font = BOLD_FONT

def build_debt_schedule_tab(ws):
    """Build the DEBT SCHEDULE tab with loan amortization"""

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 15

    row = 1

    # Title
    ws[f'A{row}'] = "DEBT SCHEDULE"
    ws[f'A{row}'].font = TITLE_FONT
    ws.merge_cells(f'A{row}:I{row}')
    row += 2

    # Summary
    ws[f'A{row}'] = "Loan Amount:"
    ws[f'B{row}'] = '=Loan_Amount'
    ws[f'B{row}'].number_format = '$#,##0'
    row += 1

    ws[f'A{row}'] = "Interest Rate:"
    ws[f'B{row}'] = '=Interest_Rate'
    ws[f'B{row}'].number_format = '0.00%'
    row += 1

    ws[f'A{row}'] = "Amortization:"
    ws[f'B{row}'] = '=Amortization_Years'
    ws[f'B{row}'].number_format = '0'
    ws[f'C{row}'] = "years"
    row += 1

    ws[f'A{row}'] = "Monthly Payment:"
    ws[f'B{row}'] = '=-PMT(Interest_Rate/12,Amortization_Years*12,Loan_Amount)'
    ws[f'B{row}'].number_format = '$#,##0'
    add_named_range(ws, f"B${row}", "Monthly_Payment")
    row += 1

    ws[f'A{row}'] = "Annual Debt Service:"
    ws[f'B{row}'] = '=Monthly_Payment*12'
    ws[f'B{row}'].number_format = '$#,##0'
    add_named_range(ws, f"B${row}", "Annual_Debt_Service")
    row += 3

    # Headers
    headers = ['Year', 'Month', 'Beg Balance', 'Principal', 'Interest', 'Total Pmt', 'End Balance', 'Cum Principal', 'Cum Interest']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row, col_idx, header)
        cell.font = BOLD_FONT
        cell.fill = INPUT_FILL
        cell.border = THIN_BORDER

    header_row = row
    row += 1

    # Monthly schedule for 360 months (30 years)
    debt_start_row = row
    for month in range(1, 361):
        year = (month - 1) // 12 + 1
        month_in_year = ((month - 1) % 12) + 1

        ws.cell(row, 1, year).number_format = '0'
        ws.cell(row, 2, month_in_year).number_format = '0'

        # Beginning balance
        if month == 1:
            ws.cell(row, 3, '=Loan_Amount')
        else:
            ws.cell(row, 3, f'=G{row-1}')
        ws.cell(row, 3).number_format = '$#,##0'

        # Interest
        ws.cell(row, 5, f'=C{row}*Interest_Rate/12')
        ws.cell(row, 5).number_format = '$#,##0'

        # Principal
        ws.cell(row, 4, f'=Monthly_Payment-E{row}')
        ws.cell(row, 4).number_format = '$#,##0'

        # Total payment
        ws.cell(row, 6, '=Monthly_Payment')
        ws.cell(row, 6).number_format = '$#,##0'

        # Ending balance
        ws.cell(row, 7, f'=C{row}-D{row}')
        ws.cell(row, 7).number_format = '$#,##0'

        # Cumulative principal
        if month == 1:
            ws.cell(row, 8, f'=D{row}')
        else:
            ws.cell(row, 8, f'=H{row-1}+D{row}')
        ws.cell(row, 8).number_format = '$#,##0'

        # Cumulative interest
        if month == 1:
            ws.cell(row, 9, f'=E{row}')
        else:
            ws.cell(row, 9, f'=I{row-1}+E{row}')
        ws.cell(row, 9).number_format = '$#,##0'

        row += 1

def build_annual_cashflow_tab(ws):
    """Build the ANNUAL CASH FLOW tab with 10-year pro forma"""

    ws.column_dimensions['A'].width = 35
    for col in range(2, 13):  # Columns B through L (Year 0-10)
        ws.column_dimensions[get_column_letter(col)].width = 13

    row = 1

    # Title
    ws[f'A{row}'] = "ANNUAL CASH FLOW PRO FORMA (10 Years)"
    ws[f'A{row}'].font = TITLE_FONT
    ws.merge_cells(f'A{row}:L{row}')
    row += 2

    # Year headers
    ws[f'A{row}'] = "Year"
    for year in range(0, 11):
        ws.cell(row, year + 2, year)
        ws.cell(row, year + 2).font = BOLD_FONT
        ws.cell(row, year + 2).fill = INPUT_FILL
        ws.cell(row, year + 2).border = THIN_BORDER
    year_header_row = row
    row += 1

    # === REVENUE ===
    ws[f'A{row}'] = "REVENUE"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL
    row += 1

    # Gross Potential Rent
    ws[f'A{row}'] = "Gross Potential Rent"
    for year in range(0, 11):
        col = year + 2
        if year == 0:
            ws.cell(row, col, '=T12_GPR')
        else:
            # Blend of in-place growth and market rent achievement
            # Simplified: Years 1-2 blend, Year 3+ full market
            if year <= 2:
                # Gradual transition to market rents
                pct_market = year * 0.4  # 40% in Y1, 80% in Y2
                formula = f'=T12_GPR*POWER(1+InPlace_Rent_Growth,{year})*(1-{pct_market})+Total_Units*Avg_Market_Rent*12*POWER(1+Market_Rent_Growth,{year})*{pct_market}'
            else:
                formula = f'=Total_Units*Avg_Market_Rent*12*POWER(1+Market_Rent_Growth,{year})'
            ws.cell(row, col, formula)
        ws.cell(row, col).number_format = '$#,##0'
    row += 1

    # Other Income
    ws[f'A{row}'] = "Other Income"
    for year in range(0, 11):
        col = year + 2
        if year == 0:
            ws.cell(row, col, '=T12_Other_Income')
        else:
            ws.cell(row, col, f'=T12_Other_Income*POWER(1+Other_Income_Growth,{year})')
        ws.cell(row, col).number_format = '$#,##0'
    row += 1

    # Gross Potential Income
    ws[f'A{row}'] = "Gross Potential Income"
    ws[f'A{row}'].font = BOLD_FONT
    gpi_row = row
    for year in range(0, 11):
        col = year + 2
        ws.cell(row, col, f'={get_column_letter(col)}{row-2}+{get_column_letter(col)}{row-1}')
        ws.cell(row, col).number_format = '$#,##0'
        ws.cell(row, col).font = BOLD_FONT
    row += 1

    # Economic Loss
    ws[f'A{row}'] = "Less: Economic Loss"
    ws[f'A{row}'].font = BOLD_FONT
    row += 1

    ws[f'A{row}'] = "  Vacancy & Credit Loss"
    for year in range(0, 11):
        col = year + 2
        # Higher vacancy during renovation (Years 0-2), then stabilized
        if year <= 2:
            vac_pct = 0.10 + (2 - year) * 0.02  # 14% Y0, 12% Y1, 10% Y2
            ws.cell(row, col, f'={get_column_letter(col)}{gpi_row}*{vac_pct}')
        else:
            ws.cell(row, col, f'={get_column_letter(col)}{gpi_row}*Stabilized_Vacancy')
        ws.cell(row, col).number_format = '$#,##0'
    row += 1

    # Effective Gross Income
    ws[f'A{row}'] = "Effective Gross Income (EGI)"
    ws[f'A{row}'].font = BOLD_FONT
    egi_row = row
    for year in range(0, 11):
        col = year + 2
        ws.cell(row, col, f'={get_column_letter(col)}{gpi_row}-{get_column_letter(col)}{row-1}')
        ws.cell(row, col).number_format = '$#,##0'
        ws.cell(row, col).font = BOLD_FONT
        ws.cell(row, col).border = Border(top=Side(style='thin'), bottom=Side(style='double'))
    row += 2

    # === OPERATING EXPENSES ===
    ws[f'A{row}'] = "OPERATING EXPENSES"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL
    row += 1

    # Property Taxes (with Prop 13)
    ws[f'A{row}'] = "Property Taxes"
    for year in range(0, 11):
        col = year + 2
        if year == 0:
            ws.cell(row, col, '=Purchase_Price*Tax_Rate+Special_Assessments')
        else:
            # Assessed value increases by Prop 13 cap (2% in CA)
            ws.cell(row, col, f'=Purchase_Price*POWER(1+Prop13_Cap,{year})*Tax_Rate+Special_Assessments')
        ws.cell(row, col).number_format = '$#,##0'
    row += 1

    # Other operating expenses (grow at OpEx growth rate)
    opex_categories = [
        ('Insurance', 'ASSUMPTIONS!B58'),
        ('Utilities', 'ASSUMPTIONS!B59+ASSUMPTIONS!B60+ASSUMPTIONS!B61+ASSUMPTIONS!B62'),
        ('Repairs & Maintenance', 'ASSUMPTIONS!B63'),
        ('Payroll', 'ASSUMPTIONS!B64'),
        ('Marketing', 'ASSUMPTIONS!B68'),
        ('Legal/Professional', 'ASSUMPTIONS!B69'),
        ('Administrative', 'ASSUMPTIONS!B70'),
    ]

    for category, t12_ref in opex_categories:
        ws[f'A{row}'] = category
        for year in range(0, 11):
            col = year + 2
            if year == 0:
                ws.cell(row, col, f'={t12_ref}')
            else:
                ws.cell(row, col, f'={t12_ref}*POWER(1+OpEx_Growth,{year})')
            ws.cell(row, col).number_format = '$#,##0'
        row += 1

    # Management Fee
    ws[f'A{row}'] = "Management Fee"
    for year in range(0, 11):
        col = year + 2
        ws.cell(row, col, f'={get_column_letter(col)}{egi_row}*Mgmt_Fee_Pct')
        ws.cell(row, col).number_format = '$#,##0'
    row += 1

    # Total Operating Expenses
    ws[f'A{row}'] = "Total Operating Expenses"
    ws[f'A{row}'].font = BOLD_FONT
    opex_row = row
    opex_start = row - 9  # Adjust based on number of categories
    for year in range(0, 11):
        col = year + 2
        ws.cell(row, col, f'=SUM({get_column_letter(col)}{opex_start}:{get_column_letter(col)}{row-1})')
        ws.cell(row, col).number_format = '$#,##0'
        ws.cell(row, col).font = BOLD_FONT
        ws.cell(row, col).border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
    row += 2

    # Net Operating Income
    ws[f'A{row}'] = "NET OPERATING INCOME (NOI)"
    ws[f'A{row}'].font = BOLD_FONT
    noi_row = row
    for year in range(0, 11):
        col = year + 2
        ws.cell(row, col, f'={get_column_letter(col)}{egi_row}-{get_column_letter(col)}{opex_row}')
        ws.cell(row, col).number_format = '$#,##0'
        ws.cell(row, col).font = BOLD_FONT
        ws.cell(row, col).border = Border(top=Side(style='double'), bottom=Side(style='double'))
    row += 1

    # NOI per Unit
    ws[f'A{row}'] = "NOI per Unit"
    for year in range(0, 11):
        col = year + 2
        ws.cell(row, col, f'={get_column_letter(col)}{noi_row}/Total_Units')
        ws.cell(row, col).number_format = '$#,##0'
    row += 1

    # NOI Margin
    ws[f'A{row}'] = "NOI Margin %"
    for year in range(0, 11):
        col = year + 2
        ws.cell(row, col, f'={get_column_letter(col)}{noi_row}/{get_column_letter(col)}{egi_row}')
        ws.cell(row, col).number_format = '0.0%'
    row += 2

    # === CAPITAL & DEBT ===
    ws[f'A{row}'] = "CAPITAL EXPENDITURES & DEBT SERVICE"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL
    row += 1

    # CapEx Reserve
    ws[f'A{row}'] = "CapEx Reserve"
    for year in range(0, 11):
        col = year + 2
        ws.cell(row, col, '=Total_Units*CapEx_Per_Unit')
        ws.cell(row, col).number_format = '$#,##0'
    capex_row = row
    row += 1

    # Renovation Expenditures
    ws[f'A{row}'] = "Renovation Expenditures"
    for year in range(0, 11):
        col = year + 2
        # Deploy renovation budget in Years 0-2
        if year == 0:
            ws.cell(row, col, '=Total_Renovation_Budget*0.4')  # 40% upfront
        elif year == 1:
            ws.cell(row, col, '=Total_Renovation_Budget*0.4')  # 40% Year 1
        elif year == 2:
            ws.cell(row, col, '=Total_Renovation_Budget*0.2')  # 20% Year 2
        else:
            ws.cell(row, col, 0)
        ws.cell(row, col).number_format = '$#,##0'
    reno_row = row
    row += 1

    # Annual Debt Service
    ws[f'A{row}'] = "Annual Debt Service"
    for year in range(0, 11):
        col = year + 2
        if year == 0:
            ws.cell(row, col, 0)  # No debt service in Year 0 (acquisition year)
        else:
            ws.cell(row, col, '=Annual_Debt_Service')
        ws.cell(row, col).number_format = '$#,##0'
    ds_row = row
    row += 1

    # DSCR
    ws[f'A{row}'] = "Debt Service Coverage Ratio"
    for year in range(0, 11):
        col = year + 2
        if year == 0:
            ws.cell(row, col, 'N/A')
        else:
            ws.cell(row, col, f'={get_column_letter(col)}{noi_row}/{get_column_letter(col)}{ds_row}')
            ws.cell(row, col).number_format = '0.00x'
    row += 2

    # Net Cash Flow
    ws[f'A{row}'] = "NET CASH FLOW (Before Sale)"
    ws[f'A{row}'].font = BOLD_FONT
    for year in range(0, 11):
        col = year + 2
        ws.cell(row, col, f'={get_column_letter(col)}{noi_row}-{get_column_letter(col)}{capex_row}-{get_column_letter(col)}{reno_row}-{get_column_letter(col)}{ds_row}')
        ws.cell(row, col).number_format = '$#,##0'
        ws.cell(row, col).font = BOLD_FONT
    ncf_row = row
    row += 2

    # === SALE PROCEEDS ===
    ws[f'A{row}'] = "SALE PROCEEDS (at exit)"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL
    row += 1

    ws[f'A{row}'] = "Forward NOI (Year N+1)"
    for year in range(0, 11):
        col = year + 2
        # Forward NOI is next year's NOI
        if year < 10:
            ws.cell(row, col, f'={get_column_letter(col+1)}{noi_row}')
        else:
            # For Year 10, project Year 11 NOI
            ws.cell(row, col, f'={get_column_letter(col)}{noi_row}*1.03')
        ws.cell(row, col).number_format = '$#,##0'
    fwd_noi_row = row
    row += 1

    ws[f'A{row}'] = "Gross Sale Price (at Exit Cap)"
    for year in range(0, 11):
        col = year + 2
        ws.cell(row, col, f'={get_column_letter(col)}{fwd_noi_row}/Exit_Cap_Rate')
        ws.cell(row, col).number_format = '$#,##0'
    gross_sale_row = row
    row += 1

    ws[f'A{row}'] = "Less: Sale Costs"
    for year in range(0, 11):
        col = year + 2
        ws.cell(row, col, f'={get_column_letter(col)}{gross_sale_row}*Sale_Costs_Pct')
        ws.cell(row, col).number_format = '$#,##0'
    sale_costs_row = row
    row += 1

    ws[f'A{row}'] = "Net Sale Proceeds (before debt)"
    for year in range(0, 11):
        col = year + 2
        ws.cell(row, col, f'={get_column_letter(col)}{gross_sale_row}-{get_column_letter(col)}{sale_costs_row}')
        ws.cell(row, col).number_format = '$#,##0'
    net_sale_row = row
    row += 1

    ws[f'A{row}'] = "Less: Loan Payoff"
    for year in range(0, 11):
        col = year + 2
        # Lookup ending balance from debt schedule
        # Year X corresponds to month X*12 in debt schedule
        debt_month = year * 12
        if debt_month > 0 and debt_month <= 360:
            ws.cell(row, col, f'="DEBT SCHEDULE"!G{9+debt_month}')  # Adjust row offset
        else:
            ws.cell(row, col, '=Loan_Amount')
        ws.cell(row, col).number_format = '$#,##0'
    loan_payoff_row = row
    row += 1

    ws[f'A{row}'] = "Net Proceeds to Equity"
    ws[f'A{row}'].font = BOLD_FONT
    for year in range(0, 11):
        col = year + 2
        ws.cell(row, col, f'={get_column_letter(col)}{net_sale_row}-{get_column_letter(col)}{loan_payoff_row}')
        ws.cell(row, col).number_format = '$#,##0'
        ws.cell(row, col).font = BOLD_FONT
    sale_proceeds_row = row
    row += 2

    # === TOTAL CASH FLOW ===
    ws[f'A{row}'] = "TOTAL CASH FLOW TO EQUITY"
    ws[f'A{row}'].font = BOLD_FONT
    for year in range(0, 11):
        col = year + 2
        if year == 0:
            # Year 0: -Equity investment + NCF
            ws.cell(row, col, f'=-Total_Equity+{get_column_letter(col)}{ncf_row}')
        else:
            # Check if this is the hold period year for sale proceeds
            ws.cell(row, col, f'=IF({year}=Hold_Period,{get_column_letter(col)}{ncf_row}+{get_column_letter(col)}{sale_proceeds_row},{get_column_letter(col)}{ncf_row})')
        ws.cell(row, col).number_format = '$#,##0'
        ws.cell(row, col).font = BOLD_FONT
        ws.cell(row, col).border = Border(top=Side(style='double'), bottom=Side(style='double'))
    total_cf_row = row
    row += 3

    # === RETURN METRICS ===
    ws[f'A{row}'] = "RETURN METRICS"
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = POSITIVE_FILL
    ws.merge_cells(f'A{row}:L{row}')
    row += 1

    ws[f'A{row}'] = "Total Equity Invested"
    ws[f'B{row}'] = '=Total_Equity'
    ws[f'B{row}'].number_format = '$#,##0'
    row += 1

    ws[f'A{row}'] = "Levered IRR"
    ws[f'A{row}'].font = BOLD_FONT
    # Use XIRR with dates or simplified IRR
    # Simplified: use IRR function on cash flows for hold period
    ws[f'B{row}'] = f'=IRR(B{total_cf_row}:L{total_cf_row})'
    ws[f'B{row}'].number_format = '0.0%'
    ws[f'B{row}'].font = BOLD_FONT
    ws[f'B{row}'].fill = POSITIVE_FILL
    add_named_range(ws, f"B${row}", "Levered_IRR")
    row += 1

    ws[f'A{row}'] = "Equity Multiple"
    ws[f'A{row}'].font = BOLD_FONT
    # Sum of all positive cash flows / equity invested
    ws[f'B{row}'] = f'=(SUM(B{total_cf_row}:L{total_cf_row})+Total_Equity)/Total_Equity'
    ws[f'B{row}'].number_format = '0.00x'
    ws[f'B{row}'].font = BOLD_FONT
    ws[f'B{row}'].fill = POSITIVE_FILL
    row += 1

    ws[f'A{row}'] = "Average Cash-on-Cash Return (Yr 3-5)"
    # Average of years 3-5 NCF / equity
    ws[f'B{row}'] = f'=AVERAGE(E{ncf_row}:G{ncf_row})/Total_Equity'
    ws[f'B{row}'].number_format = '0.0%'
    row += 1

    ws[f'A{row}'] = "Year 5 NOI"
    ws[f'B{row}'] = f'=G{noi_row}'
    ws[f'B{row}'].number_format = '$#,##0'
    row += 1

    ws[f'A{row}'] = "Yield on Cost (Stabilized NOI / Total Cost)"
    ws[f'B{row}'] = f'=G{noi_row}/Total_Uses'
    ws[f'B{row}'].number_format = '0.0%'

def build_reference_tab(ws):
    """Build the REFERENCE DATA tab"""

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15

    row = 1

    ws[f'A{row}'] = "REFERENCE DATA"
    ws[f'A{row}'].font = TITLE_FONT
    row += 2

    ws[f'A{row}'] = "This tab contains lookup data and reference tables"
    row += 2

    # Unit types
    ws[f'A{row}'] = "Unit Types"
    ws[f'A{row}'].font = BOLD_FONT
    row += 1

    unit_types = ['Studio', '1BR/1BA', '2BR/1BA', '2BR/2BA', '3BR/2BA', '3BR/2.5BA']
    for ut in unit_types:
        ws[f'A{row}'] = ut
        row += 1

    row += 1

    # Loan types
    ws[f'A{row}'] = "Loan Types"
    ws[f'A{row}'].font = BOLD_FONT
    row += 1

    loan_types = ['Agency Fixed', 'Bridge Floating', 'Bridge to Agency']
    for lt in loan_types:
        ws[f'A{row}'] = lt
        row += 1

def main():
    """Main function to build and save the workbook"""
    print("Building Aequitas Multifamily Underwriting Model...")

    wb = create_underwriting_model()

    output_path = "/Users/Helmy_LCollins/Desktop/Aequitas_Multifamily_Underwriting_Model_v1.0.xlsx"
    wb.save(output_path)

    print(f"✓ Excel model created successfully!")
    print(f"✓ Saved to: {output_path}")
    print(f"✓ Tabs included: {', '.join([ws.title for ws in wb.worksheets])}")
    print(f"\nPhase 1 (MVP) Complete:")
    print("  ✓ ASSUMPTIONS tab - all user inputs with named ranges")
    print("  ✓ SOURCES & USES tab - acquisition and renovation budget")
    print("  ✓ DEBT SCHEDULE tab - 30-year loan amortization")
    print("  ✓ ANNUAL CASH FLOW tab - 10-year pro forma")
    print("  ✓ Return calculations - IRR, equity multiple, cash-on-cash")
    print("  ✓ Sample 100-unit property loaded with realistic assumptions")
    print("\nModel is ready for testing and use!")

if __name__ == "__main__":
    main()
